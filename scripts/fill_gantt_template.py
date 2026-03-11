"""
fill_gantt_template.py  (v2)
Fills the day-bar grid in docs/Gantt-Chart-Template.xlsx with coloured bars.
Rows that already have start/end dates use those; every other task row gets
a date from the SCHEDULE dictionary below.
Saves as docs/Gantt-Chart-Template-Filled.xlsx
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import date, timedelta, datetime

# ── Paths ────────────────────────────────────────────────────────────────────
TEMPLATE_PATH = r"c:\Users\patrick josh\Desktop\BukSU\Capstone\CMS V2\docs\Gantt-Chart-Template.xlsx"
OUTPUT_PATH   = r"c:\Users\patrick josh\Desktop\BukSU\Capstone\CMS V2\docs\Gantt-Chart-Template-Filled.xlsx"
SHEET_NAME    = "Simple Gantt Chart - EX"

# ── Calendar ─────────────────────────────────────────────────────────────────
# Week 1 Monday = Feb 9 2026
WEEK1_START    = date(2026, 2, 9)
DAY_COL_START  = 10          # Column J (1-indexed openpyxl column)
TOTAL_DAY_COLS = 60          # 12 weeks x 5 days = cols 10-69 (J-BQ)
DATA_ROW_START = 11
DATA_ROW_END   = 122

# ── Fallback schedule: row -> (start_iso, end_iso) ───────────────────────────
# Applied only when the template cell has no date value.
# Dates chosen to map neatly onto the 12-week (Feb 9 - May 1 2026) grid.
SCHEDULE = {
    # PHASE 1: Data Gathering (Week 1-3, Feb 9 - Feb 27)
    13: ("2026-02-11", "2026-02-13"),
    14: ("2026-02-11", "2026-02-13"),
    15: ("2026-02-11", "2026-02-14"),
    17: ("2026-02-11", "2026-02-12"),
    18: ("2026-02-12", "2026-02-13"),
    19: ("2026-02-13", "2026-02-16"),
    21: ("2026-02-16", "2026-02-17"),
    22: ("2026-02-17", "2026-02-18"),
    23: ("2026-02-18", "2026-02-19"),
    25: ("2026-02-11", "2026-02-13"),
    26: ("2026-02-10", "2026-02-12"),
    27: ("2026-02-10", "2026-02-12"),

    # PHASE 2: System Design (Week 4-6, Mar 2 - Mar 20)
    # 30, 31 already have dates in template
    33: ("2026-02-18", "2026-02-20"),
    34: ("2026-02-18", "2026-02-20"),
    35: ("2026-02-19", "2026-02-20"),
    # 37 already has date in template
    38: ("2026-02-16", "2026-02-20"),
    39: ("2026-02-16", "2026-02-20"),
    41: ("2026-02-23", "2026-02-27"),
    42: ("2026-02-23", "2026-02-27"),
    43: ("2026-02-23", "2026-02-27"),
    45: ("2026-02-23", "2026-02-27"),
    46: ("2026-03-02", "2026-03-04"),
    47: ("2026-03-02", "2026-03-04"),
    48: ("2026-03-05", "2026-03-06"),
    50: ("2026-03-02", "2026-03-06"),
    51: ("2026-03-02", "2026-03-06"),
    52: ("2026-03-09", "2026-03-11"),
    53: ("2026-03-09", "2026-03-11"),
    55: ("2026-03-09", "2026-03-13"),
    56: ("2026-03-09", "2026-03-13"),
    58: ("2026-03-09", "2026-03-13"),
    59: ("2026-03-09", "2026-03-13"),
    60: ("2026-03-16", "2026-03-18"),
    62: ("2026-03-16", "2026-03-20"),
    63: ("2026-03-16", "2026-03-20"),
    64: ("2026-03-16", "2026-03-20"),
    66: ("2026-03-16", "2026-03-20"),

    # PHASE 3: Development (Week 7-9, Mar 23 - Apr 10)
    69: ("2026-03-23", "2026-03-27"),
    70: ("2026-03-23", "2026-03-27"),
    71: ("2026-03-23", "2026-03-27"),
    73: ("2026-03-30", "2026-04-01"),
    74: ("2026-03-30", "2026-04-01"),
    75: ("2026-03-30", "2026-04-03"),
    77: ("2026-03-30", "2026-04-02"),
    78: ("2026-04-01", "2026-04-03"),
    79: ("2026-04-01", "2026-04-03"),
    81: ("2026-04-06", "2026-04-08"),
    82: ("2026-04-06", "2026-04-08"),
    83: ("2026-04-06", "2026-04-08"),
    85: ("2026-04-08", "2026-04-10"),
    86: ("2026-04-08", "2026-04-10"),

    # PHASE 4: Testing (Week 10-11, Apr 13 - Apr 24)
    89:  ("2026-04-13", "2026-04-15"),
    90:  ("2026-04-13", "2026-04-15"),
    91:  ("2026-04-14", "2026-04-16"),
    92:  ("2026-04-15", "2026-04-17"),
    94:  ("2026-04-13", "2026-04-17"),
    95:  ("2026-04-13", "2026-04-17"),
    96:  ("2026-04-14", "2026-04-17"),
    97:  ("2026-04-14", "2026-04-17"),
    99:  ("2026-04-20", "2026-04-22"),
    100: ("2026-04-20", "2026-04-22"),
    101: ("2026-04-22", "2026-04-24"),
    103: ("2026-04-20", "2026-04-24"),
    104: ("2026-04-20", "2026-04-24"),
    105: ("2026-04-20", "2026-04-24"),

    # PHASE 5: Evaluation & Closing (Week 12, Apr 27 - May 1)
    108: ("2026-04-27", "2026-04-30"),
    109: ("2026-04-27", "2026-04-30"),
    110: ("2026-04-27", "2026-05-01"),
    112: ("2026-04-27", "2026-04-30"),
    113: ("2026-04-27", "2026-04-30"),
    115: ("2026-04-28", "2026-05-01"),
    116: ("2026-04-28", "2026-05-01"),
    117: ("2026-04-28", "2026-05-01"),
    118: ("2026-04-29", "2026-05-01"),
    120: ("2026-04-27", "2026-05-01"),
    121: ("2026-04-27", "2026-05-01"),
    122: ("2026-04-27", "2026-05-01"),
}

# ── Fill colours (AARRGGBB) ───────────────────────────────────────────────────
FILL_COMPLETE   = PatternFill("solid", fgColor="FF70AD47")  # green   - 100% done
FILL_INPROG     = PatternFill("solid", fgColor="FF4472C4")  # blue    - in progress
FILL_PLANNED    = PatternFill("solid", fgColor="FF5B9BD5")  # lt-blue - planned (0%)
FILL_ROW_BG     = PatternFill("solid", fgColor="FFDCE6F1")  # pale-bl - task row bg
FILL_SECTION_BG = PatternFill("solid", fgColor="FFD9D9D9")  # gray    - section bg


# ── Helpers ───────────────────────────────────────────────────────────────────
def to_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(v.strip(), fmt).date()
            except ValueError:
                pass
    return None


def col_for_date(d):
    """Return 1-based column index for a workday date, or None if out of grid."""
    delta = (d - WEEK1_START).days
    if delta < 0:
        return None
    week = delta // 7
    dow  = delta % 7
    if dow >= 5:           # Saturday / Sunday
        return None
    col = DAY_COL_START + week * 5 + dow
    if col >= DAY_COL_START + TOTAL_DAY_COLS:
        return None
    return col


def to_float(v):
    if v is None:
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


# ── Main ──────────────────────────────────────────────────────────────────────
print("Loading workbook ...")
wb = load_workbook(TEMPLATE_PATH)
ws = wb[SHEET_NAME]

processed  = 0
no_owner   = 0
still_empty = 0

for row in range(DATA_ROW_START, DATA_ROW_END + 1):
    owner_val = ws.cell(row=row, column=4).value   # D: Owner
    start_val = ws.cell(row=row, column=5).value   # E: Start
    due_val   = ws.cell(row=row, column=6).value   # F: End
    pct_raw   = ws.cell(row=row, column=8).value   # H: % Complete

    # Section header row (no owner) ─────────────────────────────────────────
    if not owner_val:
        for c in range(DAY_COL_START, DAY_COL_START + TOTAL_DAY_COLS):
            cell = ws.cell(row=row, column=c)
            if cell.fill.fill_type != "solid":
                cell.fill = FILL_SECTION_BG
        no_owner += 1
        continue

    # Resolve start / end dates ───────────────────────────────────────────────
    start_date = to_date(start_val)
    end_date   = to_date(due_val)

    # Use fallback schedule where template cell is blank
    if row in SCHEDULE:
        s_iso, e_iso = SCHEDULE[row]
        if not start_date:
            start_date = date.fromisoformat(s_iso)
        if not end_date:
            end_date   = date.fromisoformat(e_iso)

    if not start_date:
        still_empty += 1
        continue

    if not end_date:
        end_date = start_date + timedelta(days=4)  # default one work-week

    # Choose bar colour ────────────────────────────────────────────────────────
    pct = to_float(pct_raw)
    if pct >= 1.0:
        bar_fill = FILL_COMPLETE
    elif pct > 0.0:
        bar_fill = FILL_INPROG
    else:
        bar_fill = FILL_PLANNED

    # Paint light background across all 60 day-cols for this row
    for c in range(DAY_COL_START, DAY_COL_START + TOTAL_DAY_COLS):
        ws.cell(row=row, column=c).fill = FILL_ROW_BG

    # Paint the active bar
    cur = start_date
    while cur <= end_date:
        col_idx = col_for_date(cur)
        if col_idx is not None:
            ws.cell(row=row, column=col_idx).fill = bar_fill
        cur += timedelta(days=1)

    processed += 1

print(f"  Task rows with bars   : {processed}")
print(f"  Section header rows   : {no_owner}")
print(f"  Task rows still empty : {still_empty}")
print(f"\nSaving to:\n  {OUTPUT_PATH}")
wb.save(OUTPUT_PATH)
print("Done!")
