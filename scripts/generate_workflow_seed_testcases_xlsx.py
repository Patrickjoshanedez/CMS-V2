from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

OUT_FILE = "docs/CMS-V2-Workflow-Seed-Data-and-Test-Cases-StudentFlow.xlsx"
FONT = Font(name="Arial", size=10)
HEADER_FONT = Font(name="Arial", size=10, bold=True)
HEADER_FILL = PatternFill(fill_type="solid", start_color="D9E1F2", end_color="D9E1F2")
WRAP_ALIGN = Alignment(vertical="top", wrap_text=True)


def apply_sheet_style(ws):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = FONT
            cell.alignment = WRAP_ALIGN


def write_header(ws, headers):
    ws.append(headers)
    for idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)


def auto_width(ws, min_width=14, max_width=80):
    for column in ws.columns:
        max_len = 0
        col = column[0].column_letter
        for cell in column:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col].width = min(max(max_len + 2, min_width), max_width)


def build_credentials_sheet(wb):
    ws = wb.active
    ws.title = "Seed Credentials"
    headers = [
        "Role",
        "Email",
        "Default Password",
        "Auth Provider",
        "Active",
        "Verified",
        "Usage",
        "Source",
    ]
    write_header(ws, headers)

    rows = [
        ["instructor", "2301103203@student.buksu.edu.ph", "Password123!", "local", "Yes", "Yes", "Course/project approvals, assignment, phase advance", "server/seeders/seed-users-only.js"],
        ["adviser", "leon.mentor.buksu@gmail.com", "Password123!", "local", "Yes", "Yes", "Submission review and annotations", "server/seeders/seed-users-only.js"],
        ["panelist", "2301105311@student.buksu.edu.ph", "Password123!", "local", "Yes", "Yes", "Evaluation and defense scoring", "server/seeders/seed-users-only.js"],
        ["student (leader candidate)", "bennettchristiangeofferdon15@gmail.com", "Password123!", "local", "Yes", "Yes", "Create team and drive workflow", "server/seeders/seed-users-only.js"],
        ["student", "2301106923@student.buksu.edu.ph", "Password123!", "local", "Yes", "Yes", "Invite acceptance/member actions", "server/seeders/seed-users-only.js"],
        ["student", "chris.student.buksu@gmail.com", "Password123!", "local", "Yes", "Yes", "Invite acceptance/member actions", "server/seeders/seed-users-only.js"],
        ["student", "korjyojp@gmail.com", "Password123!", "local", "Yes", "Yes", "Invite acceptance/member actions", "server/seeders/seed-users-only.js"],
        ["student", "2501107801@student.buksu.edu.ph", "Password123!", "local", "Yes", "Yes", "Second team scenario", "server/seeders/seed-users-only.js"],
        ["student", "lara.mae.quintero@student.buksu.edu.ph", "Password123!", "local", "Yes", "Yes", "Second team scenario", "server/seeders/seed-users-only.js"],
        ["student", "student9@buksu.edu.ph", "Password123!", "local", "Yes", "Yes", "Second team scenario", "server/seeders/seed-users-only.js"],
        ["student", "student10@buksu.edu.ph", "Password123!", "local", "Yes", "Yes", "Second team scenario", "server/seeders/seed-users-only.js"],
        ["student", "student11@buksu.edu.ph", "Password123!", "local", "Yes", "Yes", "Third team draft scenario", "server/seeders/seed-users-only.js"],
        ["student", "student12@buksu.edu.ph", "Password123!", "local", "Yes", "Yes", "Third team draft scenario", "server/seeders/seed-users-only.js"],
        ["student", "student13@buksu.edu.ph", "Password123!", "local", "Yes", "Yes", "Third team draft scenario", "server/seeders/seed-users-only.js"],
        ["student (google)", "scenario.google@buksu.edu.ph", "N/A (OAuth)", "google", "Yes", "Yes", "Google-auth edge case", "server/seeders/seed-users-only.js"],
        ["student (inactive)", "scenario.inactive@buksu.edu.ph", "Password123!", "local", "No", "Yes", "Access control edge case", "server/seeders/seed-users-only.js"],
        ["student (unverified)", "scenario.unverified@buksu.edu.ph", "Password123!", "local", "Yes", "No", "Verification edge case", "server/seeders/seed-users-only.js"],
    ]

    for row in rows:
        ws.append(row)

    apply_sheet_style(ws)
    auto_width(ws)


def build_seed_prereq_sheet(wb):
    ws = wb.create_sheet("Workflow Seed Prereqs")
    headers = [
        "Stage",
        "Entity",
        "Seeded State / Required Data",
        "How to Seed / Validate",
        "Expected Outcome",
        "Source",
    ]
    write_header(ws, headers)

    rows = [
        [
            "0",
            "Database",
            "MongoDB reachable and MONGODB_URI configured",
            "Run npm run seed or node server/seeders/index.js",
            "Core users, teams, projects, submissions, evaluations loaded",
            "server/seeders/index.js",
        ],
        [
            "1",
            "Users",
            "Local users seeded with default password Password123!",
            "Run node server/seeders/seed-users-only.js",
            "You can login for all primary roles",
            "server/seeders/seed-users-only.js",
        ],
        [
            "2",
            "Current AY Teams",
            "Team Alpha locked and approved; Team Beta submitted; Team Gamma draft/no project",
            "Run full seeder (index.js)",
            "Ready-made scenarios for team, project, and phase flows",
            "server/seeders/index.js header notes",
        ],
        [
            "3",
            "Project Workflow",
            "Project statuses include draft, submitted, approved, revision-required, archived",
            "Use seeded projects + instructor account",
            "Full proposal-to-archive path testable",
            "docs/CMS-V2-Test-Cases.md",
        ],
        [
            "4",
            "Team APIs",
            "Create/invite/accept/decline/lock/list endpoints available",
            "Use /api/teams endpoints while authenticated",
            "Team creation and membership workflow validated",
            "docs/API.md",
        ],
        [
            "5",
            "Edge Cases",
            "Orphan/no-section/no-adviser/inactive/unverified/google users",
            "Login and role/state checks per account",
            "Validation and error-path coverage across auth/team flows",
            "server/seeders/seed-users-only.js",
        ],
        [
            "6",
            "Warning",
            "Full seeder clears collections before repopulation",
            "Use on intended environment only",
            "Prevent accidental data loss on non-dev DB",
            "server/seeders/index.js",
        ],
    ]

    for row in rows:
        ws.append(row)

    apply_sheet_style(ws)
    auto_width(ws)


def build_testcases_sheet(wb):
    ws = wb.create_sheet("Workflow Test Cases")
    headers = [
        "Execution Order",
        "Test Case ID",
        "Module",
        "Test Case Description",
        "Preconditions",
        "Test Steps",
        "Test Data",
        "Expected Result",
        "Role / Account",
        "Endpoint / Screen",
        "Status",
        "Remarks",
    ]
    write_header(ws, headers)

    rows = [
        [1, "TC-TEAM-001", "Team", "Create new team", "Student account with no team", "1) Login 2) Go to Teams 3) Create Team 4) Enter name 5) Save", "Team Name: Code Wizards", "Team is created and creator becomes Team Leader", "student / bennettchristiangeofferdon15@gmail.com", "POST /api/teams", "Pending", "Starts full workflow"],
        [2, "TC-TEAM-002", "Team", "Invite member to team", "Leader has an existing team", "1) Open team 2) Invite member by email 3) Send", "Invitee: 2301106923@student.buksu.edu.ph", "Invite created and pending", "student leader", "POST /api/teams/:id/invite", "Pending", "Validate max members"],
        [3, "TC-TEAM-003", "Team", "Accept team invite", "Valid invitation token exists", "1) Login as invitee 2) Open invite 3) Accept", "Invite token", "Invitee joins team", "student / 2301106923@student.buksu.edu.ph", "POST /api/teams/invites/:token/accept", "Pending", "One-team invariant"],
        [4, "TC-TEAM-010", "Team", "Prevent joining multiple teams", "User already belongs to one team", "1) Login as already-assigned student 2) Accept another invite", "Existing team + second invite", "Error shown: already in team", "student", "POST /api/teams/invites/:token/accept", "Pending", "Negative case"],
        [5, "TC-TEAM-009", "Team", "List invite candidates", "Leader logged in", "1) Open invite page", "None", "Only eligible students shown", "student leader", "Team invite UI", "Pending", "Course/section filter"],
        [6, "TC-PROJ-001", "Project", "Create new project", "Team exists with leader", "1) Login as leader 2) Create project 3) Save draft", "Title/Abstract/Keywords", "Project in draft status", "student leader", "Projects page", "Pending", "Depends on team"],
        [7, "TC-PROJ-002", "Project", "Check title similarity", "Draft project creation flow open", "1) Enter potentially similar title 2) Observe score", "Attendance Monitoring System", "Similarity warning displayed when threshold exceeded", "student leader", "Title checker", "Pending", "Similarity service"],
        [8, "TC-PROJ-003", "Project", "Submit title for approval", "Draft project exists", "1) Open project 2) Submit title", "Draft project", "Status becomes Submitted", "student leader", "Project detail", "Pending", "Locks direct edit"],
        [9, "TC-PROJ-004", "Project", "Instructor approves title", "Project in Submitted status", "1) Login instructor 2) Approve project", "Submitted project", "Status becomes Approved", "instructor", "Projects management", "Pending", "Unlocks chapter workflow"],
        [10, "TC-PROJ-009", "Project", "Assign adviser", "Project approved", "1) Login instructor 2) Assign adviser", "Adviser: leon.mentor.buksu@gmail.com", "Adviser assignment saved", "instructor", "Projects management", "Pending", "Mentor path"],
        [11, "TC-PROJ-010", "Project", "Assign panelist", "Project approved", "1) Login instructor 2) Assign panelist", "Panelist: 2301105311@student.buksu.edu.ph", "Panelist assignment saved", "instructor", "Projects management", "Pending", "Max 3 panelists"],
        [12, "TC-PROJ-013", "Project", "Set chapter deadlines", "Project approved", "1) Login instructor 2) Set deadlines", "Chapter dates", "Deadlines saved and visible to team", "instructor", "Projects management", "Pending", "Date validation"],
        [13, "TC-SUB-001", "Submission", "Upload Chapter 1", "Approved project", "1) Login leader 2) Upload chapter PDF", "chapter1.pdf", "Submission saved as Pending Review", "student leader", "Submissions page", "Pending", "PDF + size checks"],
        [14, "TC-SUB-008", "Submission", "Adviser requests revision", "Pending submission exists", "1) Login adviser 2) Open submission 3) Request revision", "Feedback text", "Revision Required status", "adviser", "Review queue", "Pending", "Feedback loop"],
        [15, "TC-SUB-009", "Submission", "Student resubmits revised file", "Revision Required status", "1) Login leader 2) Upload revised file", "chapter1-v2.pdf", "New version created and status back to Pending Review", "student leader", "Submissions page", "Pending", "Version history"],
        [16, "TC-SUB-006", "Submission", "Compile proposal (Ch1-3)", "Ch1-3 available", "1) Compile proposal 2) Upload combined PDF", "proposal.pdf", "Proposal milestone created", "student leader", "Submissions page", "Pending", "Capstone 1 milestone"],
        [17, "TC-PLAG-001", "Plagiarism", "Trigger plagiarism check", "Submission uploaded", "1) Login adviser 2) Trigger check", "Submission ID", "Job queued and processing", "adviser", "Submission detail", "Pending", "BullMQ background job"],
        [18, "TC-PLAG-003", "Plagiarism", "View plagiarism report", "Check complete", "1) Open report", "Completed report", "Originality score and match spans shown", "adviser/student leader", "Plagiarism report page", "Pending", "Threshold based decision"],
        [19, "TC-DOC-002", "Documents", "Upload manuscript to Google Docs", "Project exists", "1) Login leader 2) Upload manuscript", "manuscript.docx", "Google Docs link generated", "student leader", "Documents page", "Pending", "Drive integration"],
        [20, "TC-DOC-004", "Documents", "Sync Google Docs permissions", "Manuscript exists", "1) Sync permissions", "None", "Team + adviser have access", "student leader", "Documents page", "Pending", "Collaboration check"],
        [21, "TC-EVAL-001", "Evaluation", "Create evaluation form", "Panelist assigned to project", "1) Login panelist 2) Open assigned project 3) Evaluate", "Defense type: Midterm", "Rubric form opens", "panelist", "Evaluation page", "Pending", "Role-restricted"],
        [22, "TC-EVAL-003", "Evaluation", "Submit evaluation", "Draft evaluation completed", "1) Submit evaluation", "Completed rubric", "Evaluation locked as submitted", "panelist", "Evaluation page", "Pending", "No edit after submit"],
        [23, "TC-EVAL-005", "Evaluation", "Release evaluations to students", "Panelist evaluations submitted", "1) Login instructor 2) Release results", "Project with submitted evaluations", "Students can view results", "instructor", "Project evaluations", "Pending", "One-way release"],
        [24, "TC-PROJ-014", "Project", "Advance project phase", "Milestone complete", "1) Login instructor 2) Advance phase", "Capstone 1 complete", "Phase increases to next level", "instructor", "Projects management", "Pending", "1->2->3->4"],
        [25, "TC-SUB-015", "Submission", "Upload final academic paper", "Project in capstone phase 4", "1) Login leader 2) Upload final academic PDF", "final-academic.pdf", "Final academic doc saved", "student leader", "Final submissions", "Pending", "Required for archive"],
        [26, "TC-SUB-016", "Submission", "Upload final journal version", "Project in capstone phase 4", "1) Upload final journal PDF", "final-journal.pdf", "Final journal doc saved", "student leader", "Final submissions", "Pending", "Required for archive"],
        [27, "TC-PROJ-015", "Project", "Archive completed project", "Capstone 4 and finals complete", "1) Login instructor 2) Archive project", "Completed project", "Project appears in archive", "instructor", "Projects management", "Pending", "Archive gate checks"],
        [28, "TC-ARCH-001", "Archive", "Search archive by keyword", "Archived projects exist", "1) Search archive", "machine learning", "Matching archived projects returned", "instructor/student", "Archive page", "Pending", "Index/search validation"],
        [29, "TC-ARCH-004", "Archive", "View archived project details", "Archived project exists", "1) Open archived project", "Archived project ID", "Read-only project details visible", "any authorized user", "Archive page", "Pending", "Historical record"],
        [30, "TC-PROJ-017", "Project", "Download completion certificate", "Certificate uploaded for archived project", "1) Login team member 2) Download certificate", "Archived project", "Certificate PDF downloads", "team member", "Archived project view", "Pending", "Signed URL behavior"],
    ]

    for row in rows:
        ws.append(row)

    apply_sheet_style(ws)
    auto_width(ws)


def build_student_workflow_sheet(wb):
    ws = wb.create_sheet("Student Workflow Cases")
    headers = [
        "Flow Order",
        "Test Case ID",
        "Student Journey Stage",
        "Scenario",
        "Preconditions",
        "Steps",
        "Expected Result",
        "Primary API / Route",
        "Validation / Rule from Code",
        "Priority",
        "Status",
    ]
    write_header(ws, headers)

    rows = [
        [1, "SWF-001", "Profile Configuration", "Open Profile page and view current account data", "Authenticated student session", "1) Login 2) Open /profile 3) Load profile card", "Student profile is fetched and rendered", "GET /api/users/me | /profile", "user.routes.js exposes self-service profile route for any authenticated user", "P1", "Pending"],
        [2, "SWF-002", "Profile Configuration", "Update profile fields", "Authenticated student", "1) Edit profile values 2) Save", "Profile is updated and persisted", "PATCH /api/users/me", "updateMe route validates updateProfileSchema", "P1", "Pending"],
        [3, "SWF-003", "Profile Configuration", "Upload valid avatar image", "Authenticated student", "1) Choose avatar image 2) Upload", "Avatar upload succeeds", "POST /api/users/me/avatar", "uploadLimiter + avatarUpload + validateAvatarFile middleware", "P2", "Pending"],
        [4, "SWF-004", "Team Creation", "Create team with valid academic year", "Student has no team", "1) Go /teams 2) Create team 3) Submit", "Team created and student becomes leader", "POST /api/teams", "createTeamSchema enforces academicYear format YYYY-YYYY", "P1", "Pending"],
        [5, "SWF-005", "Team Creation", "Create team with invalid academicYear format", "Student has no team", "1) Submit form with malformed academicYear", "Validation error returned", "POST /api/teams", "Zod regex requires YYYY-YYYY", "P1", "Pending"],
        [6, "SWF-006", "Team Management", "Search invite candidates", "Leader already has team", "1) Open invite dialog 2) Search students", "Candidate list returned with canInvite flags", "GET /api/teams/:id/invite-candidates", "inviteCandidatesQuerySchema limit max 20 and search max 100", "P2", "Pending"],
        [7, "SWF-007", "Team Management", "Invite student member by email", "Leader has capacity", "1) Enter invitee email 2) Send invite", "Invite record created and notification sent", "POST /api/teams/:id/invite", "inviteMemberSchema enforces valid email; leader-only checks in team.service", "P1", "Pending"],
        [8, "SWF-008", "Team Management", "Invite duplicate/pending invite member", "Pending invite already exists", "1) Re-send same invite", "Conflict or duplicate invite error", "POST /api/teams/:id/invite", "team.service checks existing pending invite and throws DUPLICATE_INVITE", "P1", "Pending"],
        [9, "SWF-009", "Team Management", "Accept invite by token/code", "Student has valid pending invite", "1) Open accept URL 2) Confirm", "Invite accepted and team membership updated", "POST /api/teams/invites/:token/accept", "acceptInvite checks token/inviteCode and invite validity", "P1", "Pending"],
        [10, "SWF-010", "Team Management", "Decline invite", "Student has valid pending invite", "1) Open decline URL 2) Confirm", "Invite marked declined", "POST /api/teams/invites/:token/decline", "declineInvite requires valid invite + matching user email", "P2", "Pending"],
        [11, "SWF-011", "Team Finalization", "Lock team after members complete", "Requester is student leader", "1) Click finalize/lock team", "Team transitions to locked state", "PATCH /api/teams/:id/lock", "team.service returns TEAM_ALREADY_LOCKED on repeated lock", "P1", "Pending"],
        [12, "SWF-012", "Project Proposal", "Create project draft as team leader", "Student is team leader", "1) Open /project/create 2) Fill fields 3) Save", "Project is created in draft", "POST /api/projects", "project.routes guards with authorize(student) and createProjectSchema", "P1", "Pending"],
        [13, "SWF-013", "Project Proposal", "Run title similarity check", "Authenticated student", "1) Enter candidate title 2) Trigger check", "Similarity score/feedback returned", "POST /api/projects/title-check", "checkTitleSimilaritySchema validation", "P1", "Pending"],
        [14, "SWF-014", "Project Proposal", "Submit title proposal", "Project in draft", "1) Open project 2) Submit title", "Title status becomes submitted", "POST /api/projects/:id/title/submit", "submitTitleSchema validation and audit event", "P1", "Pending"],
        [15, "SWF-015", "Project Proposal", "Revise and resubmit rejected title", "Title marked revision-required", "1) Edit title/abstract 2) Resubmit", "Title returns to submitted state", "PATCH /api/projects/:id/title/revise", "checkTitleLock + updateTitleSchema", "P1", "Pending"],
        [16, "SWF-016", "Project Proposal", "Request title modification after approval", "Title already approved", "1) Submit modification request with reason", "Modification request is recorded", "POST /api/projects/:id/title/modification", "requestTitleModificationSchema requires proposed fields", "P2", "Pending"],
        [17, "SWF-017", "Submission Pipeline", "Upload chapter PDF", "Approved project with student access", "1) Open /project/submissions/upload 2) Upload file with chapter", "Submission saved for selected chapter", "POST /api/submissions/:projectId/chapters", "validatePdfFile + uploadChapterSchema require file and chapter", "P1", "Pending"],
        [18, "SWF-018", "Submission Pipeline", "Compile proposal document", "Chapters 1-3 approved/locked", "1) Upload compiled proposal PDF", "Proposal submission created", "POST /api/submissions/:projectId/proposal", "Route comment states chapters 1-3 must be locked", "P1", "Pending"],
        [19, "SWF-019", "Submission Pipeline", "View latest chapter and version history", "Project has chapter uploads", "1) Open chapter history 2) Open latest", "History and latest submission returned", "GET /api/submissions/project/:projectId/chapters/:chapter and /latest", "projectChapterParamSchema validates project/chapter params", "P2", "Pending"],
        [20, "SWF-020", "Review Feedback", "Student replies to adviser annotation thread", "Submission has annotations", "1) Open submission review 2) Reply to annotation", "Reply persisted in thread", "POST /api/submissions/:submissionId/annotations/:annotationId/replies", "Route allows student/adviser/instructor for replies", "P2", "Pending"],
        [21, "SWF-021", "Review Feedback", "Student checks feedback timeline", "Submission exists", "1) Open feedback panel", "Feedback context/timeline is displayed", "GET /api/submissions/:submissionId/feedback", "Controller route provides annotations + deadlines + review status", "P2", "Pending"],
        [22, "SWF-022", "Final Phase", "Upload final academic paper (Capstone 4)", "Project advanced to Capstone 4", "1) Upload final academic PDF", "Academic final document saved", "POST /api/submissions/:projectId/final-academic", "finalPaperSchema + validatePdfFile", "P1", "Pending"],
        [23, "SWF-023", "Final Phase", "Upload final journal paper (Capstone 4)", "Project advanced to Capstone 4", "1) Upload final journal PDF", "Journal final document saved", "POST /api/submissions/:projectId/final-journal", "finalPaperSchema + validatePdfFile", "P1", "Pending"],
        [24, "SWF-024", "Archive Access", "Search archive for prior projects", "Authenticated student", "1) Open /archive 2) Search keyword", "Archive results are returned", "GET /api/projects/archive/search | /archive", "searchArchiveQuerySchema validates query params", "P3", "Pending"],
        [25, "SWF-025", "Notifications", "Read all student notifications", "Student has unread notifications", "1) Open notifications 2) Mark all read", "Unread count becomes zero", "PATCH /api/notifications/read-all", "notification routes authenticated for all users", "P3", "Pending"],
    ]

    for row in rows:
        ws.append(row)

    apply_sheet_style(ws)
    auto_width(ws)


def main():
    wb = Workbook()
    build_credentials_sheet(wb)
    build_seed_prereq_sheet(wb)
    build_testcases_sheet(wb)
    build_student_workflow_sheet(wb)
    wb.save(OUT_FILE)
    print(f"Generated: {OUT_FILE}")


if __name__ == "__main__":
    main()
