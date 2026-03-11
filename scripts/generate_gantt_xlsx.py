from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


START_TIMELINE = date(2026, 3, 9)   # Week 1 Monday
WEEKS = 8
WORK_DAYS_PER_WEEK = 5
TOTAL_DAY_COLUMNS = WEEKS * WORK_DAYS_PER_WEEK

OUTPUT_PATH = Path(__file__).resolve().parents[1] / "Overall_Gantt_8Weeks.xlsx"

# =============================================================================
# TASK ROWS - SYSTEM DEVELOPMENT ONLY
# Format: (ID, Title, Owner, start_date, due_date, duration_days, pct_complete)
# Set Owner to None for section-header rows (dark band, no bar).
# 
# FILTERED TO INCLUDE ONLY:
#   - Architecture & Design (ARCH)
#   - Infrastructure Setup (INFRA)
#   - Backend Foundation (BE-FOUND)
#   - Authentication (AUTH)
#   - Notification System (NOTIF)
#   - Frontend Core Shell (FE-SHELL)
#   - Integration & Security (INT)
#   - Testing & QA (TEST)
#   - Deployment (DEPLOY)
#
# EXCLUDED:
#   - Planning & Research
#   - Team, Project, Submission modules (business logic)
#   - Review, Plagiarism, Evaluation modules (business workflows)
#   - Frontend workflow pages (student/adviser/panelist/instructor dashboards)
#   - Split-screen viewer
# =============================================================================
TASKS = [

    # =========================================================================
    # SECTION 2 -- ARCHITECTURE & SYSTEM DESIGN  (Week 2-3 | Mar 16-27)
    # =========================================================================
    (None,       "SECTION 2 — ARCHITECTURE & SYSTEM DESIGN", None, None, None, 0, 0),
    ("ARCH-01",  "Finalize SRS and requirements baseline document",                   "BA / Architect",    date(2026, 3, 16), date(2026, 3, 18), 3, 100),
    ("ARCH-02",  "Objective 2 – architectural pattern selection and justification",   "System Architect",  date(2026, 3, 16), date(2026, 3, 17), 2, 100),
    ("ARCH-03",  "System context diagram (Level 0 DFD) and actor boundary",          "System Architect",  date(2026, 3, 17), date(2026, 3, 17), 1, 100),
    ("ARCH-04",  "Level 1 DFD – data flows across all CMS subsystems",               "System Architect",  date(2026, 3, 17), date(2026, 3, 18), 2, 100),
    ("ARCH-05",  "Entity Relationship Diagram (ERD) – all MongoDB collections",      "Backend Lead",      date(2026, 3, 18), date(2026, 3, 19), 2, 100),
    ("ARCH-06",  "MongoDB schema design – User, Team, OTP, RefreshToken collections","Backend Lead",      date(2026, 3, 18), date(2026, 3, 19), 2, 100),
    ("ARCH-07",  "MongoDB schema design – Project, Submission, Notification, Audit", "Backend Lead",      date(2026, 3, 19), date(2026, 3, 20), 2, 100),
    ("ARCH-08",  "MongoDB indexing strategy (compound, TTL, text indexes)",           "Backend Lead",      date(2026, 3, 20), date(2026, 3, 20), 1,  80),
    ("ARCH-09",  "Use case diagram (all 4 roles – complete system scope)",            "System Architect",  date(2026, 3, 23), date(2026, 3, 24), 2, 100),
    ("ARCH-10",  "Sequence diagrams – auth, title submission, document upload flows", "System Architect",  date(2026, 3, 24), date(2026, 3, 25), 2,  80),
    ("ARCH-11",  "REST API contract – auth, user, team endpoints (OpenAPI spec)",     "Backend Lead",      date(2026, 3, 23), date(2026, 3, 24), 2,  80),
    ("ARCH-12",  "REST API contract – project, submission, review endpoints",         "Backend Lead",      date(2026, 3, 24), date(2026, 3, 25), 2,  70),
    ("ARCH-13", "REST API contract – plagiarism, evaluation, archive endpoints",       "Backend Lead",     date(2026, 3, 25), date(2026, 3, 26), 2, 100),
    ("ARCH-14",  "Backend feature-module decomposition and service boundary diagram", "System Architect",  date(2026, 3, 25), date(2026, 3, 26), 2,  80),
    ("ARCH-15",  "UI wireframes – Student dashboard and all workflow pages",          "UI/UX",             date(2026, 3, 23), date(2026, 3, 25), 3,  90),
    ("ARCH-16",  "UI wireframes – Adviser and Panelist dashboard pages",              "UI/UX",             date(2026, 3, 25), date(2026, 3, 26), 2,  90),
    ("ARCH-17",  "UI wireframes – Instructor admin dashboard and report pages",       "UI/UX",             date(2026, 3, 26), date(2026, 3, 27), 2,  80),
    ("ARCH-18",  "Component hierarchy and role-based route navigation map",           "UI/UX",             date(2026, 3, 26), date(2026, 3, 27), 2,  70),
    ("ARCH-19",  "Third-party integration contracts (S3, Copyleaks, Redis, GDrive)",  "System Architect",  date(2026, 3, 25), date(2026, 3, 26), 2,  70),
    ("ARCH-20",  "Security threat model and OWASP mitigation matrix",                 "Security Lead",     date(2026, 3, 27), date(2026, 3, 27), 1,  60),

    # =========================================================================
    # SECTION 3 -- INFRASTRUCTURE SETUP  (Week 3-4 | Mar 23 – Apr 3)
    # =========================================================================
    (None,        "SECTION 3 — INFRASTRUCTURE SETUP", None, None, None, 0, 0),
    ("INFRA-01",  "Docker Compose dev environment – service linking verification",    "DevOps",            date(2026, 3, 23), date(2026, 3, 24), 2, 100),
    ("INFRA-02",  "Environment variable schema and .env.example secrets management",  "DevOps",            date(2026, 3, 24), date(2026, 3, 24), 1, 100),
    ("INFRA-03",  "MongoDB Atlas cluster setup, connection string and replica verify","DevOps",            date(2026, 3, 24), date(2026, 3, 25), 2, 100),
    ("INFRA-04",  "Redis server config and BullMQ queue/worker initialization",       "DevOps / Backend",  date(2026, 3, 25), date(2026, 3, 26), 2, 100),
    ("INFRA-05",  "AWS S3 bucket creation, IAM policy and CORS configuration",        "DevOps",            date(2026, 3, 25), date(2026, 3, 26), 2, 100),
    ("INFRA-06",  "AWS pre-signed upload/download URL flow verification",             "DevOps / Backend",  date(2026, 3, 26), date(2026, 3, 27), 2,  90),
    ("INFRA-07", "Custom plagiarism engine (FastAPI + ChromaDB + Celery) setup",      "Backend Integ.",   date(2026, 3, 26), date(2026, 3, 27), 2, 100),
    ("INFRA-08",  "Transactional email (Mailtrap dev / SendGrid prod) configuration", "DevOps / Backend",  date(2026, 3, 27), date(2026, 3, 27), 1,  80),
    ("INFRA-09",  "ESLint, Prettier, Husky and lint-staged pre-commit hooks",         "DevOps",            date(2026, 3, 23), date(2026, 3, 24), 2, 100),
    ("INFRA-10",  "Git branching model enforcement (main/develop/feature/* rules)",   "DevOps",            date(2026, 3, 24), date(2026, 3, 24), 1, 100),
    ("INFRA-11",  "GitHub Actions CI pipeline (lint + test on every PR)",             "DevOps",            date(2026, 3, 26), date(2026, 3, 27), 2,  70),
    ("INFRA-12",  "Monorepo npm workspace dependency audit and security patch",       "DevOps",            date(2026, 3, 27), date(2026, 3, 27), 1,  70),
    ("INFRA-13",  "Nginx reverse proxy config (dev + prod) and CORS header rules",    "DevOps",            date(2026, 3, 30), date(2026, 3, 31), 2,  60),
    ("INFRA-14",  "Google Drive service account credentials and API scope setup",     "Backend Integ.",    date(2026, 3, 30), date(2026, 4,  1), 3,  50),

    # =========================================================================
    # SECTION 4 -- BACKEND FOUNDATION  (Week 4 | Mar 30 – Apr 3)
    # =========================================================================
    (None,          "SECTION 4 — BACKEND FOUNDATION", None, None, None, 0, 0),
    ("BE-FOUND-01", "Express server baseline – helmet, cors, cookieParser, bodyParser","Backend",          date(2026, 3, 30), date(2026, 3, 31), 2, 100),
    ("BE-FOUND-02", "Mongoose connection with retry logic and graceful shutdown",      "Backend",          date(2026, 3, 30), date(2026, 3, 30), 1, 100),
    ("BE-FOUND-03", "Centralized error handler (AppError class + catchAsync wrapper)", "Backend",          date(2026, 3, 30), date(2026, 3, 31), 2, 100),
    ("BE-FOUND-04", "Request validation middleware (Joi schemas per route group)",     "Backend",          date(2026, 3, 31), date(2026, 4,  1), 2,  80),
    ("BE-FOUND-05", "Audit log middleware and AuditLog Mongoose model",               "Backend",          date(2026, 3, 31), date(2026, 4,  1), 2,  80),
    ("BE-FOUND-06", "Rate limiter per route group (auth: 10/15 min, API: 100/15 min)","Backend / Sec.",   date(2026, 4,  1), date(2026, 4,  1), 1,  70),
    ("BE-FOUND-07", "File MIME magic-bytes validation middleware (server-side strict)","Backend / Sec.",   date(2026, 4,  1), date(2026, 4,  2), 2,  70),
    ("BE-FOUND-08", "BullMQ queue setup (plagiarism-check + email-dispatch workers)", "Backend / Infra",  date(2026, 4,  1), date(2026, 4,  2), 2, 100),
    ("BE-FOUND-09", "Email BullMQ worker (Nodemailer, retry policy, dead-letter queue)","Backend / Infra", date(2026, 4,  2), date(2026, 4,  3), 2,  50),
    ("BE-FOUND-10", "Socket.IO base setup, namespace config and auth handshake",      "Backend / Infra",  date(2026, 4,  2), date(2026, 4,  3), 2,  40),
    ("BE-FOUND-11", "Storage service (S3 wrapper – upload, download, delete, presign)","Backend / Infra", date(2026, 4,  2), date(2026, 4,  3), 2,  50),

    # =========================================================================
    # SECTION 5 -- AUTHENTICATION & USER MODULE  (Week 4-5 | Mar 30 – Apr 9)
    # =========================================================================
    (None,       "SECTION 5 — AUTHENTICATION & USER MODULE", None, None, None, 0, 0),
    ("AUTH-01",  "User Mongoose model (schema, indexes, static/instance methods)",    "Backend",          date(2026, 3, 30), date(2026, 3, 31), 2,  90),
    ("AUTH-02",  "OTP model (TTL index 10-min expiry, 6-digit secure generator)",     "Backend",          date(2026, 3, 31), date(2026, 3, 31), 1,  90),
    ("AUTH-03",  "RefreshToken model with rotation tracking and revocation list",     "Backend",          date(2026, 4,  1), date(2026, 4,  1), 1,  80),
    ("AUTH-04",  "Register endpoint – hash password, create user, dispatch OTP email","Backend",         date(2026, 4,  1), date(2026, 4,  2), 2,  70),
    ("AUTH-05",  "OTP verify and account activation endpoint",                        "Backend",          date(2026, 4,  2), date(2026, 4,  2), 1,  60),
    ("AUTH-06",  "Login endpoint (bcrypt compare, JWT access token + refresh cookie)","Backend",          date(2026, 4,  2), date(2026, 4,  3), 2,  60),
    ("AUTH-07",  "JWT authenticate middleware (extract, verify, attach user to req)", "Backend",          date(2026, 4,  3), date(2026, 4,  3), 1,  60),
    ("AUTH-08",  "Refresh token rotation endpoint (revoke old pair, issue new pair)", "Backend",          date(2026, 4,  3), date(2026, 4,  4), 2,  50),
    ("AUTH-09",  "Logout endpoint (clear cookies, server-side refresh revocation)",   "Backend",          date(2026, 4,  4), date(2026, 4,  4), 1,  50),
    ("AUTH-10",  "RBAC authorize middleware (role-hierarchy, route-level guards)",    "Backend",          date(2026, 4,  4), date(2026, 4,  7), 4,  50),
    ("AUTH-11",  "Password reset flow (request OTP -> verify -> set new password)",   "Backend",          date(2026, 4,  4), date(2026, 4,  7), 4,  40),
    ("AUTH-12",  "Get / update user profile endpoint with avatar S3 upload",          "Backend",          date(2026, 4,  6), date(2026, 4,  7), 2,  30),
    ("AUTH-13",  "Instructor: create/list/update role/deactivate user (admin CRUD)",  "Backend",          date(2026, 4,  7), date(2026, 4,  9), 3,  20),

    # =========================================================================
    # SECTION 11 — NOTIFICATION MODULE  (Week 5-7 | Apr 8-24)
    # =========================================================================
    (None,        "SECTION 11 — NOTIFICATION MODULE", None, None, None, 0, 0),
    ("NOTIF-01",  "Notification model (recipient, type, metadata, read status, ts)",  "Backend",          date(2026, 4,  8), date(2026, 4,  8), 1,  10),
    ("NOTIF-02",  "Notification service helper (createNotification used by all modules)","Backend",       date(2026, 4,  8), date(2026, 4,  9), 2,   5),
    ("NOTIF-03",  "Get notifications endpoint – paginated with unread count",         "Backend",          date(2026, 4,  9), date(2026, 4,  9), 1,   5),
    ("NOTIF-04",  "Mark notification read / mark-all-read endpoint",                  "Backend",          date(2026, 4,  9), date(2026, 4, 10), 2,   0),
    ("NOTIF-05",  "Socket.IO real-time notification push (emit on notif create event)","Backend",         date(2026, 4, 11), date(2026, 4, 12), 2,   0),
    ("NOTIF-06",  "Email notification templates (upload, approval, rejection, due)",  "Backend",          date(2026, 4, 12), date(2026, 4, 14), 3,   0),
    ("NOTIF-07",  "Deadline proximity cron job (24h / 48h reminder auto-dispatch)",   "Backend",          date(2026, 4, 21), date(2026, 4, 22), 2,   0),

    # =========================================================================
    # SECTION 12 -- FRONTEND CORE SHELL  (Week 4-5 | Mar 30 – Apr 10)
    # =========================================================================
    (None,           "SECTION 12 — FRONTEND CORE SHELL", None, None, None, 0, 0),
    ("FE-SHELL-01",  "Vite + React 18 config audit and workspace dependency resolution","Frontend",       date(2026, 3, 30), date(2026, 3, 30), 1, 100),
    ("FE-SHELL-02",  "Tailwind CSS configuration with strict tw- prefix enforcement", "Frontend",         date(2026, 3, 30), date(2026, 3, 31), 2, 100),
    ("FE-SHELL-03",  "React Router v6 route structure (public, protected, role-guard)","Frontend",        date(2026, 3, 31), date(2026, 4,  1), 2,  80),
    ("FE-SHELL-04",  "Axios instance with interceptors (token refresh + error norm.)", "Frontend",        date(2026, 4,  1), date(2026, 4,  2), 2,  80),
    ("FE-SHELL-05",  "Zustand auth store (user, token, role, isAuthenticated state)", "Frontend",         date(2026, 4,  2), date(2026, 4,  3), 2,  70),
    ("FE-SHELL-06",  "React Query provider and global config (staleTime, retry, gc)", "Frontend",         date(2026, 4,  3), date(2026, 4,  3), 1,  70),
    ("FE-SHELL-07",  "Dark/light mode toggle with OS preference detection",           "Frontend",         date(2026, 4,  3), date(2026, 4,  4), 2,  80),
    ("FE-SHELL-08",  "shadcn/ui component library integration and brand theming",     "Frontend",         date(2026, 4,  4), date(2026, 4,  7), 4,  70),
    ("FE-SHELL-09",  "Global error boundary component and fallback UI",               "Frontend",         date(2026, 4,  7), date(2026, 4,  8), 2,  60),
    ("FE-SHELL-10",  "Role-aware layout components (Student, Adviser, Panelist, Instr.)","Frontend",      date(2026, 4,  8), date(2026, 4, 10), 3,  50),
    ("FE-SHELL-11",  "Shared UI atoms: Button, Input, Badge, Toast, Modal, Spinner",  "Frontend",         date(2026, 4,  9), date(2026, 4, 10), 2,  50),

    # =========================================================================
    # SECTION 19 — INTEGRATION & SECURITY HARDENING  (Week 7-8 | Apr 24-30)
    # =========================================================================
    (None,       "SECTION 19 — INTEGRATION & SECURITY HARDENING", None, None, None, 0, 0),
    ("INT-01",   "E2E auth flow (register -> OTP -> login -> refresh -> logout)",     "Integration",      date(2026, 4, 24), date(2026, 4, 24), 1,   0),
    ("INT-02",   "E2E team flow (create -> invite -> accept -> lock roster)",         "Integration",      date(2026, 4, 24), date(2026, 4, 25), 2,   0),
    ("INT-03",   "E2E submission flow (upload -> plagiarism -> review -> approve)",   "Integration",      date(2026, 4, 25), date(2026, 4, 26), 2,   0),
    ("INT-04",   "Cross-module API stabilization and OpenAPI contract verification",  "Integration",      date(2026, 4, 25), date(2026, 4, 26), 2,   0),
    ("INT-05",   "Security hardening audit (OWASP top-10, injection, XSS, CSRF)",    "Sec. / Backend",   date(2026, 4, 26), date(2026, 4, 27), 2,   0),
    ("INT-06",   "API rate limiter stress test – auth/upload endpoint tuning",        "Sec. / Backend",   date(2026, 4, 27), date(2026, 4, 27), 1,   0),
    ("INT-07",   "Redis caching layer for hot reads (session, project queries)",      "Backend",          date(2026, 4, 27), date(2026, 4, 28), 2,   0),
    ("INT-08",   "Pagination and cursor-based query optimization on all list routes", "Backend",          date(2026, 4, 27), date(2026, 4, 28), 2,   0),
    ("INT-09",   "Accessibility audit (ARIA labels, keyboard nav, color contrast)",   "Frontend / QA",    date(2026, 4, 28), date(2026, 4, 29), 2,   0),
    ("INT-10",   "Responsive design audit (mobile, tablet, desktop breakpoints)",     "Frontend / QA",    date(2026, 4, 28), date(2026, 4, 29), 2,   0),

    # =========================================================================
    # SECTION 20 -- TESTING & QA  (Week 7-8 | Apr 24 – May 1)
    # =========================================================================
    (None,        "SECTION 20 — TESTING & QA", None, None, None, 0, 0),
    ("TEST-01",   "Unit tests – auth service (register, login, OTP, JWT, refresh)",   "QA / Dev",         date(2026, 4, 24), date(2026, 4, 25), 2,   0),
    ("TEST-02",   "Unit tests – team service (create, invite, accept, lock, orphan)", "QA / Dev",         date(2026, 4, 24), date(2026, 4, 25), 2,   0),
    ("TEST-03",   "Unit tests – project service (title, similarity, phase advance)",  "QA / Dev",         date(2026, 4, 25), date(2026, 4, 26), 2,   0),
    ("TEST-04",   "Unit tests – submission service (upload, versioning, locking)",    "QA / Dev",         date(2026, 4, 25), date(2026, 4, 26), 2,   0),
    ("TEST-05",   "Unit tests – plagiarism worker (mock Copyleaks, score-map)",       "QA / Dev",         date(2026, 4, 26), date(2026, 4, 27), 2,   0),
    ("TEST-06",   "Integration tests – auth routes (403/401 RBAC assertions)",        "QA",               date(2026, 4, 26), date(2026, 4, 27), 2,   0),
    ("TEST-07",   "Integration tests – team and project routes",                      "QA",               date(2026, 4, 27), date(2026, 4, 28), 2,   0),
    ("TEST-08",   "Integration tests – submission, review, and evaluation routes",    "QA",               date(2026, 4, 27), date(2026, 4, 28), 2,   0),
    ("TEST-09",   "End-to-end UI tests (Playwright – full happy path per role)",      "QA",               date(2026, 4, 28), date(2026, 4, 30), 3,   0),
    ("TEST-10",   "Bug fixing and regression cycle (all critical and high defects)",  "Dev Team",         date(2026, 4, 29), date(2026, 4, 30), 2,   0),

    # =========================================================================
    # SECTION 21 -- DEPLOYMENT & DOCUMENTATION  (Week 8 | Apr 27 – May 1)
    # =========================================================================
    (None,          "SECTION 21 — DEPLOYMENT & DOCUMENTATION", None, None, None, 0, 0),
    ("DEPLOY-01",   "Docker Compose production config (multi-stage Dockerfile build)","DevOps",           date(2026, 4, 27), date(2026, 4, 28), 2,   0),
    ("DEPLOY-02",   "Nginx production config, SSL termination and LAN routing",       "DevOps",           date(2026, 4, 28), date(2026, 4, 28), 1,   0),
    ("DEPLOY-03",   "Production environment variables and secrets rotation procedure","DevOps",           date(2026, 4, 28), date(2026, 4, 29), 2,   0),
    ("DEPLOY-04",   "Database backup, restore and point-in-time recovery procedure",  "DevOps",           date(2026, 4, 29), date(2026, 4, 29), 1,   0),
    ("DEPLOY-05",   "Deployment runbook and rollback procedure document",             "DevOps / PM",      date(2026, 4, 29), date(2026, 4, 30), 2,   0),
]

# Store for later (not displayed in system development chart):
#
# EXCLUDED BUSINESS LOGIC SECTIONS:
#
#   SECTION 1: PROJECT PLANNING & RESEARCH (PLAN-01 to PLAN-07)
#   SECTION 6: TEAM MODULE (TEAM-01 to TEAM-08)
#   SECTION 7: PROJECT & TITLE MODULE (PROJ-01 to PROJ-08)
#   SECTION 8: DOCUMENT SUBMISSION MODULE (SUB-01 to SUB-11)
#   SECTION 9: REVIEW & PLAGIARISM MODULE (REV-01 to REV-09)
#   SECTION 10: EVALUATION MODULE (EVAL-01 to EVAL-08)
#   SECTION 13: FRONTEND AUTH PAGES (FE-AUTH-01 to FE-AUTH-07)
#   SECTION 14: FRONTEND STUDENT WORKFLOW PAGES (FE-STU-01 to FE-STU-10)
#   SECTION 15: FRONTEND ADVISER PAGES (FE-ADV-01 to FE-ADV-07)
#   SECTION 16: FRONTEND PANELIST PAGES (FE-PAN-01 to FE-PAN-06)
#   SECTION 17: FRONTEND INSTRUCTOR PAGES (FE-INS-01 to FE-INS-07)
#   SECTION 18: SPLIT-SCREEN DOCUMENT VIEWER (FE-SPLIT-01 to FE-SPLIT-03)
#   SECTION 21 (Partial): Research & Documentation (DEPLOY-06, DEPLOY-07, DEPLOY-08, DEPLOY-09)
#
# To view business logic tasks separately, use: generate_gantt_xlsx_business_workflows.py


# =============================================================================
# HELPERS
# =============================================================================

def business_days(start_date: date, end_date: date) -> list[date]:
    current = start_date
    days = []
    while current <= end_date:
        if current.weekday() < 5:
            days.append(current)
        current += timedelta(days=1)
    return days


def timeline_days(start: date, weeks: int) -> list[date]:
    days = []
    current = start
    while len(days) < weeks * 5:
        if current.weekday() < 5:
            days.append(current)
        current += timedelta(days=1)
    return days


# =============================================================================
# WORKBOOK BUILD
# =============================================================================

wb = Workbook()
ws = wb.active
ws.title = "Overall Gantt"

font_header  = Font(name="Arial", bold=True, size=10)
font_section = Font(name="Arial", bold=True, size=10, color="FFFFFF")
font_body    = Font(name="Arial", size=9)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

thin   = Side(style="thin",   color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

fill_header   = PatternFill("solid", fgColor="2F3B5C")
fill_week     = PatternFill("solid", fgColor="4472C4")
fill_section  = PatternFill("solid", fgColor="1F3864")
fill_bar      = PatternFill("solid", fgColor="4F81BD")
fill_bar_done = PatternFill("solid", fgColor="70AD47")
fill_bar_prog = PatternFill("solid", fgColor="FFC000")
fill_alt_row  = PatternFill("solid", fgColor="EEF3FB")

# -- Row 1: column headers + week bands
base_headers = ["TASK ID", "TASK TITLE", "OWNER", "START", "DUE", "DAYS", "% DONE", ""]
for col, value in enumerate(base_headers, start=1):
    cell = ws.cell(row=1, column=col, value=value)
    cell.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    cell.alignment = center
    cell.fill      = fill_header
    cell.border    = border

for week in range(1, WEEKS + 1):
    start_col = 8 + (week - 1) * WORK_DAYS_PER_WEEK + 1
    end_col   = start_col + WORK_DAYS_PER_WEEK - 1
    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
    cell = ws.cell(row=1, column=start_col, value=f"WEEK {week}")
    cell.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    cell.alignment = center
    cell.fill      = fill_week
    cell.border    = border

# -- Row 2: day labels
for col in range(1, 9):
    cell = ws.cell(row=2, column=col)
    cell.font   = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    cell.fill   = fill_header
    cell.border = border

day_labels = ["M", "T", "W", "R", "F"] * WEEKS
for idx, label in enumerate(day_labels, start=9):
    cell = ws.cell(row=2, column=idx, value=label)
    cell.font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    cell.alignment = center
    cell.fill      = fill_week
    cell.border    = border

# -- Build date -> column map
project_days = timeline_days(START_TIMELINE, WEEKS)
day_to_col   = {d: 9 + i for i, d in enumerate(project_days)}

# -- Data rows
start_row  = 3
data_row   = start_row
task_count = 0

for task in TASKS:
    task_id, title, owner, start_date, due_date, duration, pct = task
    row = data_row

    if owner is None:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=1, value=f"  {title}")
        cell.font      = font_section
        cell.alignment = left
        cell.fill      = fill_section
        cell.border    = border
        for col in range(9, 9 + TOTAL_DAY_COLUMNS):
            c = ws.cell(row=row, column=col)
            c.fill   = fill_section
            c.border = border
        ws.row_dimensions[row].height = 16
    else:
        row_fill = fill_alt_row if task_count % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        values = [task_id, title, owner,
                  start_date.isoformat(), due_date.isoformat(),
                  duration, pct / 100, ""]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font      = font_body
            cell.alignment = left if col in (2, 3) else center
            cell.fill      = row_fill
            cell.border    = border

        ws.cell(row=row, column=7).number_format = "0%"

        if pct >= 100:
            bar_fill = fill_bar_done
        elif pct > 0:
            bar_fill = fill_bar_prog
        else:
            bar_fill = fill_bar

        active_days = set(business_days(start_date, due_date))
        for day in project_days:
            col  = day_to_col[day]
            cell = ws.cell(row=row, column=col)
            cell.fill   = bar_fill if day in active_days else row_fill
            cell.border = border

        ws.row_dimensions[row].height = 15
        task_count += 1

    data_row += 1

# -- Column widths
widths = {1: 13, 2: 62, 3: 22, 4: 12, 5: 12, 6: 7, 7: 8, 8: 2}
for col, width in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width
for col in range(9, 9 + TOTAL_DAY_COLUMNS):
    ws.column_dimensions[get_column_letter(col)].width = 3.2

ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 18

ws.freeze_panes = "I3"

# -- Legend
legend_row = data_row + 1
ws.cell(row=legend_row, column=1, value="LEGEND").font = font_header
legend_items = [
    (fill_bar_done, "Completed (100%)"),
    (fill_bar_prog, "In Progress (> 0%)"),
    (fill_bar,      "Planned (0%)"),
    (fill_section,  "Section Header"),
]
for i, (lfill, ltext) in enumerate(legend_items):
    r = legend_row + 1 + i
    ws.cell(row=r, column=1).fill   = lfill
    ws.cell(row=r, column=1).border = border
    lc = ws.cell(row=r, column=2, value=ltext)
    lc.font      = font_body
    lc.alignment = left
    lc.border    = border

wb.save(OUTPUT_PATH)
print(f"Saved  -> {OUTPUT_PATH}")
print(f"Task rows   : {task_count}")
print(f"Total rows  : {data_row - start_row} (includes {data_row - start_row - task_count} section headers)")
