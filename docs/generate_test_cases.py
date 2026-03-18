"""
Generate comprehensive functional & non-functional test cases for CMS V2.
Outputs: docs/CMS_V2_Test_Cases.xlsx
"""
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Styles ────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
MODULE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

COLUMNS = [
    ("Test Case ID", 14),
    ("Module", 18),
    ("Sub-Module", 20),
    ("Test Scenario", 40),
    ("Pre-Conditions", 35),
    ("Test Steps", 50),
    ("Expected Result", 40),
    ("Priority", 10),
    ("Status", 10),
    ("Remarks", 25),
]

def make_sheet(ws, title: str, test_cases: list[list[object]]) -> None:
    """Populate a worksheet with the standard header and test cases."""
    ws.title = title
    ws.sheet_properties.tabColor = "2F5496"

    # Header row
    for ci, (col_name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.row_dimensions[1].height = 30
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    ws.freeze_panes = "A2"

    for ri, tc in enumerate(test_cases, 2):
        for ci, val in enumerate(tc, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = WRAP
            cell.border = THIN_BORDER
        ws.row_dimensions[ri].height = 55


# ══════════════════════════════════════════════════════════════════
# FUNCTIONAL TEST CASES
# ══════════════════════════════════════════════════════════════════

functional: list[list[object]] = []
tc_id = 1

def tc(module: str, sub: str, scenario: str, pre: str, steps: str, expected: str, priority: str = "High") -> None:
    global tc_id
    row = [f"FTC-{tc_id:03d}", module, sub, scenario, pre, steps, expected, priority, "", ""]
    functional.append(row)
    tc_id += 1

# ─── 1. AUTH MODULE ───────────────────────────────────────────────

tc("Authentication", "Registration",
   "Register with valid data",
   "User is on the registration page; reCAPTCHA loaded",
   "1. Fill firstName, lastName, email, password (Abc12345)\n2. Complete reCAPTCHA\n3. Click Register",
   "Account created. OTP sent to email. Redirect to OTP verification page.")

tc("Authentication", "Registration",
   "Register with duplicate email",
   "An account with test@buksu.edu.ph already exists",
   "1. Enter existing email test@buksu.edu.ph\n2. Fill other fields\n3. Click Register",
   "Error message: 'Email already in use'. No duplicate account created.")

tc("Authentication", "Registration",
   "Register with weak password",
   "User is on the registration page",
   "1. Enter password 'abc' (no uppercase, no digit, too short)\n2. Submit",
   "Validation error: password must be 8+ chars with uppercase, lowercase, and digit.")

tc("Authentication", "OTP Verification",
   "Verify with valid OTP",
   "User registered; OTP email received",
   "1. Enter the 6-digit OTP from email\n2. Click Verify",
   "Account verified. isVerified set to true. Redirect to login page.")

tc("Authentication", "OTP Verification",
   "Verify with expired OTP",
   "OTP was sent > 10 minutes ago",
   "1. Enter the expired OTP\n2. Click Verify",
   "Error: 'OTP expired'. Prompt to resend.")

tc("Authentication", "OTP Verification",
   "Resend OTP",
   "User on OTP page; previous OTP expired",
   "1. Click Resend OTP\n2. Check email",
   "New OTP sent. Previous OTP invalidated. Rate limited to 3 req/10 min.")

tc("Authentication", "Login",
   "Login with valid credentials",
   "Verified account exists",
   "1. Enter email and password\n2. Complete reCAPTCHA\n3. Click Login",
   "JWT access token set in HTTP-only cookie. Redirect to dashboard. lastLoginAt updated.")

tc("Authentication", "Login",
   "Login with invalid password",
   "Account exists with different password",
   "1. Enter correct email, wrong password\n2. Click Login",
   "Error: 'Invalid credentials'. No token issued.")

tc("Authentication", "Login",
   "Login with unverified account",
   "Account registered but OTP not verified",
   "1. Enter valid credentials\n2. Click Login",
   "Error: 'Account not verified'. Redirect to OTP page.")

tc("Authentication", "Login",
   "Login with deactivated account",
   "Account isActive = false",
   "1. Enter valid credentials\n2. Click Login",
   "Error: 'Account deactivated'. Contact administrator.")

tc("Authentication", "Google OAuth",
   "Login with Google account (first time)",
   "Google OAuth configured; user has Google account",
   "1. Click 'Sign in with Google'\n2. Select Google account\n3. Grant permissions",
   "New account auto-created. authProvider set to 'google'. Redirect to dashboard.")

tc("Authentication", "Google OAuth",
   "Login with Google account (existing local)",
   "Local account exists with same email",
   "1. Click 'Sign in with Google'\n2. Select matching Google account",
   "Google ID linked to existing account. Login successful.")

tc("Authentication", "Token Refresh",
   "Refresh expired access token",
   "User logged in; access token expired; refresh token valid",
   "1. Make API request\n2. System auto-refreshes token",
   "New access token issued. Old refresh token revoked (rotation). Request proceeds.")

tc("Authentication", "Token Refresh",
   "Detect refresh token reuse",
   "Refresh token already used once (stolen scenario)",
   "1. Use the old/revoked refresh token",
   "All user sessions revoked. Token family invalidated. User must re-login.")

tc("Authentication", "Forgot Password",
   "Reset password via email OTP",
   "Account exists and is verified",
   "1. Click Forgot Password\n2. Enter email\n3. Receive OTP\n4. Enter OTP + new password",
   "Password updated. All refresh tokens revoked. User can login with new password.")

tc("Authentication", "Forgot Password",
   "Reset with same password as current",
   "User initiating password reset",
   "1. Enter OTP\n2. Set new password same as current",
   "Error: 'PASSWORD_REUSE — Cannot reuse current password'.")

tc("Authentication", "Change Password",
   "Change password while logged in",
   "User is authenticated",
   "1. Go to Profile\n2. Enter current password + new password\n3. Submit",
   "Password changed. All other sessions (refresh tokens) revoked.")

tc("Authentication", "Logout",
   "Logout user",
   "User is logged in",
   "1. Click Logout button",
   "Cookies cleared. Refresh tokens revoked. Redirect to login page.")

tc("Authentication", "Rate Limiting",
   "Exceed login rate limit",
   "User on login page",
   "1. Attempt login 11 times within 15 minutes",
   "429 Too Many Requests after 10th attempt. Blocked for remaining window.", "Medium")

# ─── 2. USER MANAGEMENT ──────────────────────────────────────────

tc("User Management", "Profile",
   "View own profile",
   "User is authenticated",
   "1. Navigate to Profile page\n2. Verify displayed info",
   "Profile shows firstName, lastName, middleName, email, role, avatar, section.")

tc("User Management", "Profile",
   "Update own profile",
   "User is authenticated",
   "1. Edit firstName and middleName\n2. Click Save",
   "Profile updated. Changes reflected immediately.")

tc("User Management", "Avatar Upload",
   "Upload valid avatar image",
   "User is on profile page",
   "1. Click avatar upload\n2. Select valid JPEG (< 5MB)\n3. Confirm",
   "Avatar uploaded to S3. Profile picture URL updated.")

tc("User Management", "Avatar Upload",
   "Upload invalid file as avatar",
   "User is on profile page",
   "1. Click avatar upload\n2. Select a .exe file renamed to .jpg",
   "Rejected by magic-byte validation. Error: invalid image format.")

tc("User Management", "Admin - List Users",
   "List all users with pagination",
   "Logged in as Instructor",
   "1. Go to Users page\n2. Verify pagination controls\n3. Navigate to page 2",
   "Users displayed in paginated table. Max 100 per page. Total count shown.")

tc("User Management", "Admin - Create User",
   "Create a new user account",
   "Logged in as Instructor",
   "1. Click Add User\n2. Fill name, email, password, role\n3. Submit",
   "User created with isVerified=true. Audit log entry created.")

tc("User Management", "Admin - Change Role",
   "Change user role",
   "Logged in as Instructor; target user exists",
   "1. Select user\n2. Click Change Role\n3. Select 'adviser'\n4. Confirm",
   "Role updated. Audit log records role change.")

tc("User Management", "Admin - Deactivate",
   "Deactivate a user",
   "Logged in as Instructor; target user is active",
   "1. Select user\n2. Click Deactivate\n3. Confirm",
   "User isActive=false. User cannot login. Audit log created.")

tc("User Management", "Filter/Search",
   "Search users by name",
   "Logged in as Instructor; multiple users exist",
   "1. Type 'Reyes' in search bar\n2. Press Enter",
   "Only users matching 'Reyes' shown. Search max 100 chars.")

tc("User Management", "Instructor List",
   "List all instructors",
   "Authenticated user (any role)",
   "1. Call GET /api/users/instructors",
   "Returns array of all users with role=instructor. Available for adviser assignment.")

# ─── 3. TEAM MANAGEMENT ──────────────────────────────────────────

tc("Team Management", "Create Team",
   "Create a new team",
   "Student has no team; logged in",
   "1. Click Create Team\n2. Enter team name ('Team Delta') and academic year\n3. Submit",
   "Team created. Student becomes leader. Student's teamId updated.")

tc("Team Management", "Create Team",
   "Create team when already in a team",
   "Student already belongs to a team",
   "1. Attempt to create a new team",
   "Error: 'Already in a team'. Cannot create duplicate team membership.")

tc("Team Management", "Invite Member",
   "Invite student by email",
   "Leader of an unlocked team",
   "1. Enter student email\n2. Click Invite",
   "Invite created. Email sent. Invite token valid for 48 hours.")

tc("Team Management", "Invite Member",
   "Invite to a full team (4 members)",
   "Team already has 4 members",
   "1. Enter 5th student email\n2. Click Invite",
   "Error: 'Team is full'. Max 4 members enforced.")

tc("Team Management", "Accept Invite",
   "Accept team invitation",
   "Student received invite; not in any team",
   "1. Click accept link/button\n2. Confirm",
   "Student added to team. teamId set. Invite status = accepted.")

tc("Team Management", "Decline Invite",
   "Decline team invitation",
   "Student received invite",
   "1. Click Decline",
   "Invite status = declined. Student remains without team.")

tc("Team Management", "Finalize Team",
   "Finalize team members",
   "Leader of team with members",
   "1. Click 'Finalize Team Members'\n2. Confirm dialog",
   "team.isLocked = true. No further member changes allowed. Badge shows 'Finalized'.")

tc("Team Management", "Finalize Team",
   "Attempt changes on finalized team",
   "Team is locked/finalized",
   "1. Try to invite a new member",
   "Error: 'Team is locked'. Invite not created.")

tc("Team Management", "View Team",
   "View my team details",
   "Student is in a team",
   "1. Navigate to Teams page",
   "Shows team name, members, leader badge, finalized status, academic year.")

tc("Team Management", "List Teams (Admin)",
   "List all teams with filters",
   "Logged in as Instructor",
   "1. Go to Teams page\n2. Filter by academic year\n3. Filter by locked status",
   "Filtered teams displayed. Pagination working. Shows member count.")

# ─── 4. PROJECT MANAGEMENT ───────────────────────────────────────

tc("Project Management", "Create Project",
   "Create project as team leader",
   "Team is locked; no existing project; student is leader",
   "1. Click Create Project\n2. Enter title (10+ chars), abstract, keywords\n3. Select section and assign member roles\n4. Submit",
   "Project created with titleStatus=draft. memberRoleAssignments saved. One project per team enforced.")

tc("Project Management", "Create Project",
   "Create project with team not locked",
   "Team is NOT locked",
   "1. Try to create a project",
   "Error: 'Team must be finalized first'. Project not created.")

tc("Project Management", "Title Workflow",
   "Submit title for approval",
   "Project in draft status",
   "1. Review title, abstract, keywords\n2. Click Submit Title",
   "titleStatus changed to 'submitted'. Notification sent to instructors. Title becomes read-only.")

tc("Project Management", "Title Workflow",
   "Approve submitted title (Instructor)",
   "Project titleStatus = submitted; logged in as Instructor",
   "1. View project\n2. Click Approve Title\n3. Confirm",
   "titleStatus = approved. Audit log created. Notification sent to team. Capstone 1 tab unlocks after panelist assignment.")

tc("Project Management", "Title Workflow",
   "Reject submitted title (Instructor)",
   "Project titleStatus = submitted; logged in as Instructor",
   "1. View project\n2. Click Reject\n3. Enter reason (5-1000 chars)\n4. Submit",
   "titleStatus = revision_required. Rejection reason saved. Notification sent. Student can revise & resubmit.")

tc("Project Management", "Title Workflow",
   "Revise and resubmit title",
   "titleStatus = revision_required",
   "1. Edit title/abstract/keywords\n2. Click Revise & Resubmit",
   "titleStatus changes back to 'submitted'. New version available for review.")

tc("Project Management", "Title Workflow",
   "Request title modification (post-approval)",
   "titleStatus = approved",
   "1. Click Request Title Modification\n2. Enter proposed title (10-300 chars) + justification (20-1000 chars)\n3. Submit",
   "titleModificationRequest created with status=pending. Title remains approved until resolved.")

tc("Project Management", "Title Workflow",
   "Resolve title modification (Instructor)",
   "Pending modification request exists",
   "1. View modification request\n2. Click Approve or Deny\n3. Add optional review note",
   "If approved: title updated to proposed title. If denied: request marked denied. Audit log created.")

tc("Project Management", "Title Similarity",
   "Check title similarity in real-time",
   "User typing a project title",
   "1. Enter title text (10+ chars)\n2. System auto-checks similarity",
   "Returns similarity score. Warning if score > titleSimilarityThreshold (default 0.65).")

tc("Project Management", "Adviser Assignment",
   "Assign adviser to project (Instructor)",
   "Project exists; adviser user exists",
   "1. Select project\n2. Click Assign Adviser\n3. Choose adviser\n4. Confirm",
   "adviserId set on project. Student users' instructorId updated. Notifications sent. Audit log created.")

tc("Project Management", "Panelist Assignment",
   "Assign panelist to project (Instructor)",
   "Project exists; < 3 panelists assigned",
   "1. Select project\n2. Click Add Panelist\n3. Choose panelist\n4. Confirm",
   "Panelist added to panelistIds[]. Notification sent. Max 3 panelists enforced.")

tc("Project Management", "Panelist Assignment",
   "Panelist self-selects to project",
   "Logged in as Panelist; project has < 3 panelists",
   "1. View project list\n2. Click 'Select as Panelist'",
   "Panelist's ID added to project.panelistIds[]. Audit log + notification created.")

tc("Project Management", "Panelist Assignment",
   "Remove panelist from project",
   "Instructor logged in; panelist is assigned",
   "1. View project\n2. Click Remove next to panelist name\n3. Confirm",
   "Panelist removed from panelistIds[]. Notification sent to panelist.")

tc("Project Management", "Deadlines",
   "Set chapter deadlines",
   "Instructor or Adviser; project exists",
   "1. Open project\n2. Set deadlines for chapters 1-5, proposal, defense\n3. Mark some as TBA\n4. Save",
   "Deadlines saved. TBA array updated. Notification sent to team.")

tc("Project Management", "Phase Advancement",
   "Advance capstone phase",
   "Instructor logged in; project at phase 1",
   "1. View project\n2. Click Advance Phase\n3. Confirm",
   "capstonePhase = 2. Next tab unlocks in student view. Audit log + notification created.")

tc("Project Management", "Project Rejection",
   "Reject entire project",
   "Instructor logged in; project is active",
   "1. View project\n2. Click Reject Project\n3. Enter reason (5-1000 chars)\n4. Confirm",
   "projectStatus = rejected. Rejection reason saved. Audit log created. Team notified.")

tc("Project Management", "Prototypes",
   "Upload prototype image",
   "Student; project at phase 2+; Capstone 2 tab active",
   "1. Click Add Prototype\n2. Enter title and description\n3. Upload JPEG/PNG image (< 50MB)\n4. Submit",
   "Prototype saved with type=image, storageKey set. Max 20 prototypes enforced.")

tc("Project Management", "Prototypes",
   "Add prototype link",
   "Student; project at phase 2+",
   "1. Click Add Link Prototype\n2. Enter title, description, URL\n3. Submit",
   "Prototype saved with type=link, url set. Notification sent.")

tc("Project Management", "Prototypes",
   "Delete prototype",
   "Student; prototype exists",
   "1. Click Delete on prototype\n2. Confirm",
   "Prototype removed from project.prototypes[]. File deleted from S3 if applicable.")

tc("Project Management", "Archive",
   "Archive completed project (Instructor)",
   "Project at phase 4; all evaluations complete",
   "1. View project\n2. Click Archive\n3. Enter completion notes\n4. Confirm",
   "isArchived=true, archivedAt set, projectStatus=archived. Audit log. Notification to team.")

tc("Project Management", "Certificate",
   "Upload completion certificate (Instructor)",
   "Project is archived",
   "1. View archived project\n2. Click Upload Certificate\n3. Select PDF file\n4. Upload",
   "Certificate stored in S3. certificateStorageKey set. Notification sent.")

tc("Project Management", "Certificate",
   "Download completion certificate",
   "Certificate uploaded; user is team member or faculty",
   "1. Click Download Certificate",
   "Pre-signed S3 URL generated. File downloads. 5-min expiry on URL.")

tc("Project Management", "List Projects",
   "Filter projects by status and year",
   "Instructor logged in; multiple projects exist",
   "1. Go to Projects page\n2. Filter by academicYear=2025-2026\n3. Filter by titleStatus=approved",
   "Only matching projects shown. Pagination works. Text search on title/keywords available.", "Medium")

# ─── 5. TAB-BASED WORKFLOW ───────────────────────────────────────

tc("Student Workflow", "Tab Navigation",
   "Proposal tab always accessible",
   "Student with a project",
   "1. Go to My Project page",
   "Proposal tab is always unlocked and clickable. Shows project info, title status, title actions.")

tc("Student Workflow", "Tab Navigation",
   "Capstone 1 tab locks until title approved + panelists assigned",
   "titleStatus=draft, no panelists",
   "1. View My Project page\n2. Click Capstone 1 tab",
   "Capstone 1 tab shows lock icon, is grayed out, and cannot be clicked.")

tc("Student Workflow", "Tab Navigation",
   "Capstone 1 tab unlocks when conditions met",
   "titleStatus=approved AND panelistIds.length > 0",
   "1. View My Project page",
   "Capstone 1 tab is clickable. Shows chapters 1-3, evaluation panel with defense type 'proposal'.")

tc("Student Workflow", "Tab Navigation",
   "Capstone 2 tab unlocks at phase 2",
   "Instructor advanced project to phase 2",
   "1. View My Project page",
   "Capstone 2 tab unlocked. Shows prototype upload/gallery and evaluation panel (defenseType=midterm).")

tc("Student Workflow", "Tab Navigation",
   "Capstone 3 tab unlocks at phase 3",
   "Instructor advanced project to phase 3",
   "1. View My Project page",
   "Capstone 3 tab unlocked. Shows chapters 4-5 progress and evaluation panel (defenseType=paper).")

tc("Student Workflow", "Tab Navigation",
   "Final Defense tab unlocks at phase 4",
   "Instructor advanced project to phase 4",
   "1. View My Project page",
   "Final tab unlocked. Shows final paper upload and evaluation panel (defenseType=final).")

tc("Student Workflow", "Tab Navigation",
   "Auto-advance to highest unlocked tab",
   "Project at phase 3",
   "1. Navigate to My Project page (fresh load)",
   "Tab auto-selects Capstone 3 (highest unlocked tab). User can still click previous tabs.")

tc("Student Workflow", "Proposal Tab",
   "Show panelists pending card after approval",
   "titleStatus=approved BUT no panelists assigned yet",
   "1. View Proposal tab",
   "Green 'Panelists Pending' card shown. Message: waiting for instructor to assign panelists.")

# ─── 6. SUBMISSIONS ──────────────────────────────────────────────

tc("Submissions", "Chapter Upload",
   "Upload chapter 1 document",
   "Student; project exists; chapter 1 deadline set",
   "1. Navigate to Capstone 1 tab\n2. Click Upload for Chapter 1\n3. Select PDF file\n4. Add remarks\n5. Submit",
   "Submission created with type=chapter, chapter=1, version=1. File stored in S3. Plagiarism check queued.")

tc("Submissions", "Chapter Upload",
   "Re-upload chapter (version increment)",
   "Chapter 1 v1 exists with status revisions_required",
   "1. Upload new version of chapter 1",
   "New submission with version=2 created. Links to same project+chapter. Previous version preserved.")

tc("Submissions", "Chapter Upload",
   "Upload after deadline (late flag)",
   "Chapter 1 deadline has passed",
   "1. Upload chapter 1 after deadline",
   "Submission created with isLate=true. Visible flag on submission record.")

tc("Submissions", "Proposal Compilation",
   "Compile proposal from chapters 1-3",
   "Chapters 1, 2, 3 all locked (approved)",
   "1. Navigate to compile proposal\n2. Upload compiled document\n3. Submit",
   "Proposal submission created. type=proposal. All 3 chapters must be locked first.")

tc("Submissions", "Proposal Compilation",
   "Compile proposal with unlocked chapters",
   "Chapter 2 still pending",
   "1. Attempt to compile proposal",
   "Error: 'Chapters 1-3 must be approved/locked before compiling proposal'. Upload blocked.")

tc("Submissions", "Final Papers",
   "Upload final academic paper",
   "Project at phase 4; student logged in",
   "1. Navigate to Final Defense tab\n2. Click Upload Academic Paper\n3. Select file\n4. Submit",
   "Submission created with type=final_academic. Unique per project (only one allowed).")

tc("Submissions", "Final Papers",
   "Upload journal version",
   "Project at phase 4",
   "1. Click Upload Journal\n2. Select file\n3. Submit",
   "Submission with type=final_journal created. Separate from academic version.")

tc("Submissions", "Review",
   "Approve submission (Adviser)",
   "Adviser logged in; submission status=pending or under_review",
   "1. View submission\n2. Click Approve\n3. Add review note\n4. Confirm",
   "Status = approved → locked. Review note saved. reviewedBy set. Notification to student.")

tc("Submissions", "Review",
   "Request revisions on submission",
   "Adviser logged in; submission under review",
   "1. View submission\n2. Click Revisions Required\n3. Add review note (required)\n4. Submit",
   "Status = revisions_required. Student notified to re-upload with changes.")

tc("Submissions", "Review",
   "Reject submission",
   "Adviser logged in; submission under review",
   "1. View submission\n2. Click Reject\n3. Add reason\n4. Confirm",
   "Status = rejected. Student notified. Must upload new version.")

tc("Submissions", "Unlock",
   "Unlock locked submission",
   "Submission status = locked; Adviser or Instructor logged in",
   "1. Click Unlock\n2. Enter reason (10-1000 chars)\n3. Confirm",
   "Submission unlocked. Student can re-upload. Audit log created.")

tc("Submissions", "Annotations",
   "Add annotation to submission",
   "Adviser viewing a submission PDF",
   "1. Select text/area on page\n2. Enter comment (1-2000 chars)\n3. Save",
   "Annotation saved with page number, content, highlight coordinates. Student notified.")

tc("Submissions", "Annotations",
   "Delete annotation",
   "Annotation exists; same adviser who created it",
   "1. Click delete on annotation\n2. Confirm",
   "Annotation removed from submission.")

tc("Submissions", "View",
   "View submission document",
   "Submission exists; user is authorized",
   "1. Click View on submission",
   "Pre-signed S3 URL generated (5-min expiry). PDF viewer opens.")

tc("Submissions", "Version History",
   "View chapter version history",
   "Chapter 1 has 3 versions",
   "1. Navigate to Chapter 1\n2. View version history",
   "All 3 versions listed with version number, status, date, reviewer. Latest shown by default.")

# ─── 7. PLAGIARISM ───────────────────────────────────────────────

tc("Plagiarism", "Auto Check",
   "Automatic plagiarism check on upload",
   "Chapter uploaded; plagiarism engine configured",
   "1. Upload chapter document",
   "Plagiarism check queued automatically. Status = queued → processing → completed.")

tc("Plagiarism", "Manual Check",
   "Trigger manual plagiarism check",
   "Adviser/Instructor; submission exists",
   "1. View submission\n2. Click Run Plagiarism Check",
   "New check queued. Status updates in real-time. Overwrites previous result.")

tc("Plagiarism", "Results",
   "View plagiarism report",
   "Plagiarism check completed",
   "1. Click View Plagiarism Report",
   "Shows similarity percentage, matched sources with excerpts, warning flag if > threshold.")

tc("Plagiarism", "Threshold Warning",
   "Similarity exceeds threshold",
   "Plagiarism score > plagiarismThreshold (default 75%)",
   "1. View plagiarism result",
   "Warning flag shown. Highlighted matched sources. Adviser notified.")

tc("Plagiarism", "Corpus Management",
   "Index submission in corpus",
   "Adviser/Instructor; submission exists",
   "1. Click Index in Corpus",
   "Submission text added to plagiarism corpus for future comparisons.")

tc("Plagiarism", "Corpus Management",
   "Remove from corpus",
   "Submission indexed in corpus",
   "1. Click Remove from Corpus",
   "Submission removed. Won't match in future checks.", "Medium")

# ─── 8. EVALUATIONS ──────────────────────────────────────────────

tc("Evaluations", "Get or Create",
   "Panelist opens evaluation for project",
   "Panelist assigned to project; proposal defense",
   "1. View project\n2. Navigate to evaluation panel",
   "If no draft exists: new evaluation auto-created with default criteria (4 criteria, max 3 pts each for proposal). If draft exists: show existing.")

tc("Evaluations", "Criteria Defaults",
   "Verify Capstone 1 rubric criteria",
   "Panelist opens proposal evaluation",
   "1. Check auto-generated criteria",
   "Shows: Problem Definition (max 3), Writing Quality (max 3), Originality (max 3), Independence (max 3).")

tc("Evaluations", "Criteria Defaults",
   "Verify Capstone 2 rubric criteria",
   "Panelist opens midterm evaluation",
   "1. Check auto-generated criteria",
   "Shows 8 criteria (max 4 pts each): Completeness, System Dev Progress, Alignment, Technical Quality, Documentation, Timeline, Problem ID/Resolution, Presentation.")

tc("Evaluations", "Criteria Defaults",
   "Verify Capstone 3 rubric criteria",
   "Panelist opens paper evaluation",
   "1. Check auto-generated criteria",
   "Shows 5 criteria (max 3 pts each): Presentation of Results, Analysis, Summary/Conclusions/Recs, Writing Quality, Independence.")

tc("Evaluations", "Criteria Defaults",
   "Verify Final defense rubric criteria",
   "Panelist opens final evaluation",
   "1. Check auto-generated criteria",
   "Shows 6 criteria: System Functionality (25), Technical Implementation (20), Documentation (15), Innovation (15), Presentation (15), Q&A (10).")

tc("Evaluations", "Score Entry",
   "Enter scores and save draft",
   "Panelist; draft evaluation exists",
   "1. Enter scores for each criterion\n2. Add overall comment\n3. Save as Draft",
   "Scores saved. totalScore and maxTotalScore computed. Status remains draft. Can edit later.")

tc("Evaluations", "Submit",
   "Submit evaluation (lock from edits)",
   "All criteria scored; draft evaluation",
   "1. Click Submit Evaluation\n2. Confirm",
   "Status = submitted. submittedAt recorded. Evaluation locked — no further edits. Audit log created.")

tc("Evaluations", "Release",
   "Release evaluations to students (Instructor)",
   "All panelists submitted for this defense type",
   "1. View project as Instructor\n2. Click Release Evaluations\n3. Confirm",
   "Status = released for all evals of that defense type. releasedAt set. Students can now view scores. Notification sent.")

tc("Evaluations", "Visibility",
   "Student can only see released evaluations",
   "2 panelists: one submitted (unreleased), one released",
   "1. Student views evaluation panel",
   "Only released evaluations visible. Submitted-but-not-released are hidden from student.")

tc("Evaluations", "One Per Panelist",
   "Enforce unique evaluation constraint",
   "Panelist already has evaluation for this project + defense type",
   "1. Attempt to create duplicate evaluation",
   "Existing evaluation returned. No duplicate created (unique compound index).")

# ─── 9. DOCUMENTS (Google Drive) ─────────────────────────────────

tc("Documents", "Template",
   "Create document template (Instructor)",
   "Google Docs integration configured",
   "1. Enter Google Doc ID\n2. Set title, description, document type\n3. Create",
   "Template created. Links to Google Doc. Available for student use.")

tc("Documents", "Template",
   "List templates by type",
   "Templates exist",
   "1. Filter templates by documentType (e.g., chapter_1)\n2. Filter by isActive",
   "Filtered templates displayed. Only active templates shown by default.")

tc("Documents", "Generate",
   "Generate document from template",
   "Student; template exists; project exists",
   "1. Select template\n2. Click Generate Document\n3. Open in Google Docs",
   "Google Doc copy created. ProjectDocument record saved. One doc per type per project enforced.")

tc("Documents", "Generate",
   "Attempt duplicate document type",
   "Project already has a chapter_1 document",
   "1. Try to generate another chapter_1 document",
   "Error: duplicate. Unique index on projectId + documentType prevents creation.")

tc("Documents", "Delete",
   "Delete project document",
   "Project document exists; Student or Adviser",
   "1. Click Delete Document\n2. Confirm",
   "ProjectDocument record deleted. Google Doc may still exist in Drive.", "Medium")

# ─── 10. NOTIFICATIONS ───────────────────────────────────────────

tc("Notifications", "Receive",
   "Receive notification on title approval",
   "Team member; title just approved by instructor",
   "1. Check notification bell",
   "Notification shown: 'Your project title has been approved'. type=title_approved. isRead=false.")

tc("Notifications", "Read",
   "Mark notification as read",
   "Unread notification exists",
   "1. Click on notification",
   "isRead = true. Unread count decrements.")

tc("Notifications", "Mark All Read",
   "Mark all notifications as read",
   "Multiple unread notifications",
   "1. Click 'Mark All Read'",
   "All notifications for user set isRead=true. Unread count = 0.")

tc("Notifications", "Delete",
   "Delete single notification",
   "Notification exists",
   "1. Click delete icon on notification\n2. Confirm",
   "Notification removed from database.")

tc("Notifications", "Clear All",
   "Clear all notifications",
   "Multiple notifications exist",
   "1. Click Clear All\n2. Confirm",
   "All user notifications deleted.")

tc("Notifications", "Real-time",
   "Real-time notification via WebSocket",
   "Socket.io connected; adviser submits review",
   "1. Student is on any page\n2. Adviser reviews their submission",
   "Notification appears immediately without page refresh. Bell icon count updates.", "Medium")

# ─── 11. DASHBOARD ───────────────────────────────────────────────

tc("Dashboard", "Student",
   "Student dashboard shows project overview",
   "Student with active project",
   "1. Go to Dashboard",
   "Shows team info, project status, chapter progress (1-5 with status/version), recent notifications.")

tc("Dashboard", "Instructor",
   "Instructor dashboard shows system overview",
   "Instructor logged in",
   "1. Go to Dashboard",
   "Shows: total users, teams, projects. Pending title approvals (top 10). Recent submissions. Projects-by-status breakdown.")

tc("Dashboard", "Adviser",
   "Adviser dashboard shows assigned projects",
   "Adviser with assigned projects",
   "1. Go to Dashboard",
   "Shows: assigned projects list, pending reviews count, active project count, recent notifications.")

tc("Dashboard", "Panelist",
   "Panelist dashboard shows evaluation tasks",
   "Panelist assigned to projects",
   "1. Go to Dashboard",
   "Shows: assigned projects, pending evaluations (draft status), active project count.")

# ─── 12. SETTINGS ────────────────────────────────────────────────

tc("Settings", "View",
   "View system settings",
   "Authenticated user",
   "1. Go to Settings page",
   "Shows: plagiarism threshold (0-100), title similarity threshold (0-1), max file size, system announcement.")

tc("Settings", "Update",
   "Update system settings (Instructor)",
   "Instructor logged in",
   "1. Change plagiarism threshold to 80\n2. Set announcement text\n3. Save",
   "Settings updated. Audit log created. Previous values overwritten. updatedBy recorded.")

tc("Settings", "Validation",
   "Invalid settings values",
   "Instructor",
   "1. Set plagiarism threshold to 150\n2. Submit",
   "Validation error: 'must be between 0 and 100'. No changes saved.", "Medium")

# ─── 13. AUDIT LOG ───────────────────────────────────────────────

tc("Audit Log", "View",
   "View audit trail (Instructor)",
   "Instructor logged in; actions have been performed",
   "1. Go to Audit Log page\n2. Browse entries",
   "Shows chronological list: action, actor, actor role, target, description, timestamp, IP address.")

tc("Audit Log", "Filter",
   "Filter audit logs by date range",
   "Audit logs exist across dates",
   "1. Set start date and end date\n2. Filter by action type\n3. Filter by target type",
   "Only matching logs shown. Pagination max 100 per page.")

tc("Audit Log", "Entity History",
   "View audit history for specific entity",
   "Adviser; project has audit entries",
   "1. View project\n2. Check audit/activity history",
   "Lists all actions on that entity: created, title submitted, title approved, etc.")

# ─── 14. ARCHIVE & SEARCH ────────────────────────────────────────

tc("Archive", "Search",
   "Search archived projects",
   "Archived projects exist from previous years",
   "1. Go to Archive page\n2. Enter search term 'enrollment'\n3. Filter by year 2024-2025",
   "Matching archived projects displayed. Shows title, abstract, keywords, year, team.")

tc("Archive", "Search",
   "Search by keyword",
   "Archived projects with various keywords",
   "1. Enter keyword filter 'machine learning'",
   "Only projects with matching keyword shown. Max 50 results per page.")

tc("Archive", "Bulk Upload",
   "Bulk upload legacy archive document (Instructor)",
   "Instructor on Bulk Upload page",
   "1. Enter title, abstract, keywords, academic year\n2. Attach PDF/DOCX\n3. Upload",
   "Archive record created for projects predating the system. File stored in S3.")

tc("Archive", "Reports",
   "Generate project reports (Instructor)",
   "Instructor logged in; projects exist",
   "1. Go to Reports page\n2. Select academic year\n3. Optionally filter by adviser\n4. Generate",
   "Report generated showing projects, statuses, evaluation scores, completion rates.")

# ─── 15. ACADEMICS ───────────────────────────────────────────────

tc("Academics", "Course",
   "Create a new course (Instructor)",
   "Instructor logged in",
   "1. Go to Academics/Course management\n2. Enter name and code\n3. Submit",
   "Course created. Code auto-uppercased. Unique code enforced.")

tc("Academics", "Section",
   "Create a new section",
   "Course exists; Instructor logged in",
   "1. Select course\n2. Enter section name, code, academic year\n3. Submit",
   "Section created. Unique constraint: courseId + academicYear + name.")

tc("Academics", "Section",
   "List sections filtered by course and year",
   "Sections exist",
   "1. Filter by courseId\n2. Filter by academicYear",
   "Only matching sections shown. Both active and inactive (previous years).")

tc("Academics", "Hierarchy",
   "View academic hierarchy",
   "Instructor; courses and sections exist",
   "1. Call GET /api/academics/hierarchy",
   "Nested structure: courses → sections per year. Used for dropdowns and assignments.", "Medium")


# ══════════════════════════════════════════════════════════════════
# NON-FUNCTIONAL TEST CASES
# ══════════════════════════════════════════════════════════════════

nonfunctional: list[list[object]] = []
ntc_id = 1

def ntc(module: str, sub: str, scenario: str, pre: str, steps: str, expected: str, priority: str = "High") -> None:
    global ntc_id
    row = [f"NTC-{ntc_id:03d}", module, sub, scenario, pre, steps, expected, priority, "", ""]
    nonfunctional.append(row)
    ntc_id += 1

# ─── SECURITY ────────────────────────────────────────────────────

ntc("Security", "Authentication",
    "Access protected endpoint without token",
    "No auth cookie set",
    "1. Call GET /api/users/me without cookies",
    "401 Unauthorized. No data leaked.")

ntc("Security", "Authentication",
    "Access with expired JWT token",
    "JWT access token expired (> 15 min)",
    "1. Make API request with expired token",
    "401 Unauthorized. Client should auto-refresh via /api/auth/refresh.")

ntc("Security", "Authentication",
    "Access with tampered JWT token",
    "Modify JWT payload without re-signing",
    "1. Alter userId in JWT payload\n2. Call protected endpoint",
    "401 Invalid token. jsonwebtoken verification fails.")

ntc("Security", "Authorization",
    "Student accessing instructor-only endpoint",
    "Logged in as student",
    "1. Call PUT /api/settings/ (instructor-only)",
    "403 Forbidden. Role check blocks access.")

ntc("Security", "Authorization",
    "Panelist accessing adviser-only endpoint",
    "Logged in as panelist",
    "1. Call POST /api/submissions/:id/review (adviser/instructor only)",
    "403 Forbidden.")

ntc("Security", "CSRF Protection",
    "JWT stored in HTTP-only cookies",
    "User logged in",
    "1. Inspect browser cookies via developer tools",
    "accessToken and refreshToken are HTTP-only, Secure, SameSite=Strict. Not accessible to JavaScript.")

ntc("Security", "XSS Prevention",
    "Inject script in project title",
    "Student creating a project",
    "1. Enter title: <script>alert('xss')</script>\n2. Submit",
    "Input sanitized/escaped. Script not executed. HTML entities rendered as text.")

ntc("Security", "SQL/NoSQL Injection",
    "Inject NoSQL operator in login",
    "On login page",
    "1. Set email to {\"$gt\": \"\"} in request body\n2. Submit",
    "Zod validation rejects non-string input. Mongoose parameterized queries prevent injection.")

ntc("Security", "File Upload Security",
    "Upload malicious file with renamed extension",
    "Upload submission endpoint",
    "1. Rename malware.exe to document.pdf\n2. Upload as chapter",
    "Magic-byte validation (file-type library) rejects file. Only PDF/DOCX/TXT magic signatures accepted.")

ntc("Security", "Rate Limiting",
    "Brute force login attempt",
    "Attacker script",
    "1. Send 11 login requests in 15-minute window",
    "authLimiter blocks at 10 req/15 min. Returns 429 Too Many Requests.")

ntc("Security", "Rate Limiting",
    "OTP brute force attempt",
    "Attacker targeting OTP endpoint",
    "1. Send 4 OTP resend requests in 10 minutes",
    "otpLimiter blocks at 3 req/10 min. Prevents OTP enumeration.")

ntc("Security", "Password Security",
    "Passwords stored securely",
    "User registration or password change",
    "1. Create user\n2. Inspect database directly",
    "Password stored as bcrypt hash (salt 12). select:false prevents return in queries. Original password never stored.")

ntc("Security", "Data Privacy",
    "Password not returned in API responses",
    "Authenticated user",
    "1. Call GET /api/users/me\n2. Inspect response",
    "No password or __v field in response. toJSON transform removes sensitive fields.")

ntc("Security", "Helmet Headers",
    "Security headers present",
    "Any API response",
    "1. Inspect response headers",
    "Helmet sets: X-Content-Type-Options, X-Frame-Options, Strict-Transport-Security, Content-Security-Policy, etc.")

ntc("Security", "CORS",
    "Cross-origin request from unauthorized domain",
    "Request from unauthorized origin",
    "1. Make API request from unauthorized origin",
    "CORS rejects request. Only configured origins allowed.")

ntc("Security", "CAPTCHA",
    "Login without CAPTCHA",
    "reCAPTCHA configured in production",
    "1. Submit login without captchaToken",
    "verifyCaptcha middleware rejects request. CAPTCHA required for register/login.")

# ─── PERFORMANCE ─────────────────────────────────────────────────

ntc("Performance", "Response Time",
    "API response time under normal load",
    "Server running; DB populated with seed data",
    "1. Call GET /api/dashboard/stats\n2. Measure response time",
    "Response returned within 500ms. Indexed queries ensure fast lookup.")

ntc("Performance", "Response Time",
    "File upload response time",
    "10MB PDF chapter upload",
    "1. Upload 10MB file\n2. Measure time to response",
    "Upload + S3 storage completes within 5 seconds on stable connection.", "Medium")

ntc("Performance", "Pagination",
    "Large dataset pagination performance",
    "100+ projects in database",
    "1. Call GET /api/projects?page=10&limit=10\n2. Measure response time",
    "Paginated response within 300ms. Count query efficient with indexes.")

ntc("Performance", "Database Indexes",
    "Query performance with proper indexes",
    "All model indexes configured",
    "1. Run explain() on common queries\n2. Verify index usage",
    "Queries use indexes (text indexes for search, compound for filtering). No collection scans.")

ntc("Performance", "Rate Limiter Memory",
    "Rate limiter doesn't cause memory leak",
    "Server running for extended period",
    "1. Monitor memory usage over 24 hours\n2. Simulate varied traffic",
    "Memory remains stable. Rate limiter IP entries expire. No unbounded growth.", "Low")

# ─── RELIABILITY ─────────────────────────────────────────────────

ntc("Reliability", "Error Handling",
    "Graceful error on database disconnection",
    "MongoDB temporarily unreachable",
    "1. Kill MongoDB connection\n2. Make API request",
    "500 Internal Server Error with structured JSON. No stack trace in production. App doesn't crash.")

ntc("Reliability", "Error Handling",
    "Consistent error format across all endpoints",
    "Various error triggers",
    "1. Trigger 400, 401, 403, 404, 409, 422, 500\n2. Compare response shapes",
    "All errors follow: { success: false, error: { code, message, status } }. No leaked internals.")

ntc("Reliability", "Data Integrity",
    "Unique constraints enforced",
    "Attempt duplicate data creation",
    "1. Create user with existing email\n2. Create duplicate project for team\n3. Create duplicate evaluation",
    "11000 duplicate key error. Descriptive message returned. No corrupt data.")

ntc("Reliability", "Data Integrity",
    "Mongoose validation catches invalid data",
    "Input violates schema constraints",
    "1. Submit title with < 10 chars\n2. Submit keyword array with > 10 items",
    "Validation error with field-level messages. No invalid data persisted.")

ntc("Reliability", "Audit Immutability",
    "Audit logs cannot be modified or deleted",
    "Audit log entries exist",
    "1. Attempt to modify audit entry\n2. Attempt to delete via API",
    "No update/delete endpoints exist for audit logs. Append-only by design.")

# ─── USABILITY ────────────────────────────────────────────────────

ntc("Usability", "UI Navigation",
    "Consistent navigation across pages",
    "User on any page",
    "1. Check sidebar/header navigation\n2. Navigate between pages",
    "Sidebar highlights current page. Breadcrumbs where applicable. Back navigation works.", "Medium")

ntc("Usability", "Responsiveness",
    "Mobile-responsive layout",
    "Access on mobile viewport (375px wide)",
    "1. Open app on mobile device\n2. Navigate all pages",
    "All pages readable. Tables scroll horizontally. Forms stack vertically. Touch targets adequate.", "Medium")

ntc("Usability", "Form Validation",
    "Client-side validation feedback",
    "User filling out forms",
    "1. Leave required field empty\n2. Enter invalid email\n3. Enter short password",
    "Real-time validation feedback. Error messages shown inline. Submit button disabled for invalid forms.", "Medium")

ntc("Usability", "Loading States",
    "Loading indicators on async operations",
    "User performing actions",
    "1. Upload file\n2. Submit form\n3. Navigate to data-heavy page",
    "Spinner/skeleton shown during loading. Buttons disabled to prevent double-submit.", "Medium")

ntc("Usability", "Error Messages",
    "User-friendly error messages",
    "Various error scenarios",
    "1. Trigger validation error\n2. Trigger auth error\n3. Trigger conflict error",
    "Messages in plain language. No technical jargon or stack traces. Actionable guidance provided.", "Medium")

ntc("Usability", "Tab Workflow",
    "Locked tabs give clear feedback",
    "Student with early-phase project",
    "1. View locked Capstone 2/3/Final tabs",
    "Lock icon visible. Grayed out. Cursor shows not-allowed. No confusing click behavior.", "Medium")

# ─── COMPATIBILITY ────────────────────────────────────────────────

ntc("Compatibility", "Browser",
    "Works on Chrome, Firefox, Edge",
    "App deployed; browsers available",
    "1. Open app in Chrome\n2. Open in Firefox\n3. Open in Edge\n4. Test login, navigation, uploads",
    "All features functional. No browser-specific rendering issues. Console-free of errors.", "Medium")

ntc("Compatibility", "Node.js",
    "Server runs on Node.js v18+",
    "Node.js version installed",
    "1. Check node -v\n2. Run npm start\n3. Verify no deprecation warnings",
    "Server starts without warnings on Node 18/20/22. All dependencies compatible.", "Low")

# ─── SCALABILITY ─────────────────────────────────────────────────

ntc("Scalability", "Concurrent Users",
    "Handle 50 concurrent users",
    "Load testing tool configured",
    "1. Simulate 50 users performing mixed operations\n2. Monitor response times and errors",
    "All requests served. Average response < 1s. No 500 errors. Rate limiter handles abuse.", "Medium")

ntc("Scalability", "Data Volume",
    "System handles 500+ projects",
    "Database populated with 500 projects",
    "1. Paginate through projects\n2. Search projects\n3. Generate reports",
    "Pagination responsive. Text search returns results in < 500ms. Reports generate correctly.", "Medium")

ntc("Scalability", "File Storage",
    "S3 handles growing file volume",
    "Many files uploaded over time",
    "1. Verify pre-signed URL generation remains fast\n2. Check S3 bucket size",
    "URLs generated instantly. No degradation. Storage unlimited by S3 design.", "Low")

# ─── MAINTAINABILITY ─────────────────────────────────────────────

ntc("Maintainability", "Code Quality",
    "Consistent code structure across modules",
    "Codebase review",
    "1. Compare module structures (routes, controller, service, validation, model)\n2. Check naming conventions",
    "All 13 server modules follow same pattern. Shared constants centralized. No duplicate logic.", "Low")

ntc("Maintainability", "Environment Config",
    "Environment variables properly externalized",
    "Check .env usage",
    "1. Verify all secrets in .env\n2. Verify .env.example exists\n3. Check no hardcoded credentials",
    "All sensitive config externalized. .env not committed to git. Example file documents required vars.", "Low")

ntc("Maintainability", "API Versioning Ready",
    "API routes organized for future versioning",
    "Review route structure",
    "1. Check route prefix structure\n2. Verify controller separation",
    "Routes under /api/. Controllers separated by module. Easy to version if needed.", "Low")

# ── Build workbook ────────────────────────────────────────────────

# Sheet 1: Functional
ws_func = wb.active
make_sheet(ws_func, "Functional Test Cases", functional)

# Sheet 2: Non-Functional
ws_nf = wb.create_sheet()
make_sheet(ws_nf, "Non-Functional Test Cases", nonfunctional)

# Sheet 3: Summary
ws_sum = wb.create_sheet("Summary")
summary_data = [
    ["CMS V2 — Test Case Summary", "", "", ""],
    ["", "", "", ""],
    ["Category", "Module", "Test Cases", "Priority Breakdown"],
    ["Functional", "Authentication", 18, "17 High, 1 Medium"],
    ["Functional", "User Management", 10, "10 High"],
    ["Functional", "Team Management", 10, "10 High"],
    ["Functional", "Project Management", 18, "17 High, 1 Medium"],
    ["Functional", "Student Workflow (Tabs)", 8, "8 High"],
    ["Functional", "Submissions", 14, "14 High"],
    ["Functional", "Plagiarism", 6, "5 High, 1 Medium"],
    ["Functional", "Evaluations", 10, "10 High"],
    ["Functional", "Documents (Drive)", 5, "4 High, 1 Medium"],
    ["Functional", "Notifications", 6, "5 High, 1 Medium"],
    ["Functional", "Dashboard", 4, "4 High"],
    ["Functional", "Settings", 3, "2 High, 1 Medium"],
    ["Functional", "Audit Log", 3, "3 High"],
    ["Functional", "Archive & Search", 4, "4 High"],
    ["Functional", "Academics", 4, "3 High, 1 Medium"],
    ["", "", "", ""],
    ["Non-Functional", "Security", 16, "16 High"],
    ["Non-Functional", "Performance", 5, "2 High, 2 Medium, 1 Low"],
    ["Non-Functional", "Reliability", 5, "5 High"],
    ["Non-Functional", "Usability", 6, "6 Medium"],
    ["Non-Functional", "Compatibility", 2, "1 Medium, 1 Low"],
    ["Non-Functional", "Scalability", 3, "3 Medium"],
    ["Non-Functional", "Maintainability", 3, "3 Low"],
    ["", "", "", ""],
    ["TOTAL FUNCTIONAL", "", len(functional), ""],
    ["TOTAL NON-FUNCTIONAL", "", len(nonfunctional), ""],
    ["GRAND TOTAL", "", len(functional) + len(nonfunctional), ""],
]

for ri, row in enumerate(summary_data, 1):
    for ci, val in enumerate(row, 1):
        cell = ws_sum.cell(row=ri, column=ci, value=val)
        cell.border = THIN_BORDER if ri >= 3 else Border()
        cell.alignment = WRAP
        if ri == 1:
            cell.font = Font(name="Calibri", size=14, bold=True, color="2F5496")
        elif ri == 3:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        elif ri >= len(summary_data) - 2:
            cell.font = Font(bold=True)
            cell.fill = MODULE_FILL

ws_sum.column_dimensions["A"].width = 20
ws_sum.column_dimensions["B"].width = 28
ws_sum.column_dimensions["C"].width = 14
ws_sum.column_dimensions["D"].width = 25

def main() -> None:
   output_path = Path(__file__).resolve().parent / "CMS_V2_Test_Cases.xlsx"
   wb.save(output_path)
   print(f"Test cases saved to {output_path}")
   print(f"Functional:     {len(functional)} test cases")
   print(f"Non-Functional: {len(nonfunctional)} test cases")
   print(f"Grand Total:    {len(functional) + len(nonfunctional)} test cases")


if __name__ == "__main__":
   main()
